"""
eco_playwright.py
-----------------
Automação do ECO Requisition para criação de POs via Playwright.

Parâmetro `confirmar`:
  - callable(titulo, mensagem) → True (prosseguir) / False (cancelar)
  - None = modo autônomo sem pausas (usar após testes aprovados)
"""

import asyncio
import json
import os
import re
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

import logging

from config import SHIP_VIA_MAP, GL_CODE_PLANILHA
from utils import numero_cotacao as _numero_cotacao_util, norm_vendor as _norm_vendor

logger = logging.getLogger(__name__)

ECO_URL         = "https://requisition.chouest.com"
TIMEOUT         = 30_000
SHORT_WAIT      = 5_000
VENDOR_MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor_map.json")

# Mapeamento de UOM extraída → opção no ECO Requisition (kendo-dropdownlist)
# Chaves em lowercase; valores EXATOS como aparecem no dropdown do ECO
UOM_MAP = {
    "each": "each", "ea": "each", "pc": "each", "pcs": "each", "piece": "each",
    "unit": "each", "un": "each", "und": "each",
    "box": "box", "bx": "box",
    "case": "case", "cs": "case",
    "cm": "cm",
    "cu yd": "cu yd", "cubic yard": "cu yd",
    "day": "day", "days": "day",
    "dm": "dm",
    "dozen": "dozen", "dz": "dozen", "doz": "dozen",
    "drum": "drum",
    "feet": "feet", "ft": "feet", "foot": "feet",
    "gal": "gal", "gallon": "gal", "gallons": "gal",
    "hour": "hour", "hr": "hour", "hrs": "hour",
    "lb": "lb", "lbs": "lb", "pound": "lb", "pounds": "lb",
    "liter": "liter", "ltr": "liter", "l": "liter", "litre": "liter",
    "meter": "meter", "m": "meter", "mtr": "meter", "metre": "meter",
    "miles": "miles", "mi": "miles", "mile": "miles",
    "month": "month", "mo": "month", "months": "month",
    "oz": "oz", "ounce": "oz", "ounces": "oz",
    "pack": "pack", "pk": "pack",
    "pail": "pail",
    "pair": "pair", "pr": "pair", "pairs": "pair",
    "quart": "quart", "qt": "quart",
    "set": "set", "kit": "set", "lot": "set",
    "sq ft": "sq ft", "sqft": "sq ft", "square foot": "sq ft", "square feet": "sq ft",
    "ton": "ton", "tons": "ton",
    "week": "week", "wk": "week", "weeks": "week",
}


# ─────────────────────────────────────────────────────────────────────
# Memória de fornecedores — persiste entre execuções
# ─────────────────────────────────────────────────────────────────────

def _carregar_vendor_map() -> dict:
    """
    Lê o mapa de fornecedores salvo (chave → texto_opção_ECO).
    Normaliza as chaves ao carregar para que 'k-mar', 'kmar', 'K-MAR' etc.
    resultem na mesma entrada.
    """
    try:
        with open(VENDOR_MAP_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {_norm_vendor(k): v for k, v in raw.items()}
    except Exception:
        logger.warning("Não foi possível carregar vendor_map.json — usando mapa vazio.")
        return {}


def _salvar_vendor_map(vendor_map: dict):
    """Persiste o mapa de fornecedores em disco."""
    try:
        with open(VENDOR_MAP_FILE, "w", encoding="utf-8") as f:
            json.dump(vendor_map, f, ensure_ascii=False, indent=2)
    except Exception:
        logger.warning("Não foi possível salvar vendor_map.json.")


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def _numero_cotacao(analise: dict) -> str:
    return _numero_cotacao_util(analise) or ""


def _termo_busca_vendor(nome: str) -> str:
    """
    Retorna o termo de busca para o autocomplete de vendor no ECO.
    Usa apenas as primeiras 2 palavras com 3+ caracteres para maximizar
    as chances de encontrar autocomplete (ex: 'Master Control Systems Inc'
    → 'Master Control').
    """
    if not nome:
        return ""
    palavras_sig = [w for w in nome.split() if len(w) >= 3]
    if not palavras_sig:
        return nome
    return " ".join(palavras_sig[:2])


def _carregar_vessels() -> dict:
    """
    Lê a planilha 'Brazil Vessels - GL CODE.xlsx' e retorna
    {NOME_EMBARCAÇÃO_MAIUSCULO: gl_code_str}.
    A planilha tem múltiplos grupos de colunas lado a lado:
      Grupo 1 : col 0 = nome, col 1 = GL
      Grupos N: col 3N = nome, col 3N+2 = GL  (N = 1, 2, 3, ...)
    """
    from openpyxl import load_workbook
    vessels = {}
    try:
        wb = load_workbook(GL_CODE_PLANILHA, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Grupo 1: col A = nome, col B = GL
            if row[0] and row[1]:
                vessels[str(row[0]).upper().strip()] = str(int(row[1]))
            # Grupos seguintes: col 3, 6, 9, ... = nome; col 5, 8, 11, ... = GL
            col = 3
            while col + 2 < len(row):
                nome = row[col]
                gl   = row[col + 2]
                if nome and gl:
                    vessels[str(nome).upper().strip()] = str(int(gl))
                col += 3
        wb.close()
    except Exception:
        logger.warning("Não foi possível carregar planilha de GL Codes: %s", GL_CODE_PLANILHA)
    return vessels


async def _ok(confirmar, titulo: str, mensagem: str) -> bool:
    """
    Pausa para confirmação do usuário. Se confirmar=None, retorna True direto.
    Roda o callback bloqueante em executor para não travar o event loop.
    """
    if confirmar is None:
        return True
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, confirmar, titulo, mensagem)


async def _escolher_async(escolher, titulo: str, opcoes: list) -> int:
    """
    Exibe lista de opções numeradas e retorna o índice (0-based) escolhido.
    Se escolher=None (modo autônomo), retorna 0 (primeira opção).
    Retorna -1 se o usuário cancelar.
    """
    if escolher is None:
        return 0
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, escolher, titulo, opcoes)


# ─────────────────────────────────────────────────────────────────────
# Login — confirmação em cada sub-passo
# ─────────────────────────────────────────────────────────────────────

async def _login(page, usuario: str, senha: str, confirmar):
    await page.goto(f"{ECO_URL}/login")
    await page.wait_for_selector("#username", timeout=TIMEOUT)
    await page.fill("#username", usuario)
    pw = page.locator("kendo-formfield").nth(1).locator("input").first
    await pw.fill(senha)
    await page.locator("button[type='submit']").first.click()
    await page.wait_for_url(f"{ECO_URL}/**", timeout=TIMEOUT)
    await page.wait_for_load_state("networkidle", timeout=TIMEOUT)


# ─────────────────────────────────────────────────────────────────────
# Criação de PO — confirmação em cada sub-passo
# ─────────────────────────────────────────────────────────────────────

async def _criar_po_par(page, par: dict, vessels: dict, confirmar, escolher, vendor_map: dict) -> dict:
    analise       = par["analise"]
    po_data       = analise.get("po", {})
    melhor        = analise.get("melhor_preco") or {}
    numero_cot    = _numero_cotacao(analise)
    numero_po     = po_data.get("numero_po") or "?"
    centro_custo  = (po_data.get("centro_de_custo") or po_data.get("solicitante") or "").strip()
    forn_extraido = (po_data.get("fornecedor_escolhido_comentario") or "").strip()
    # NÃO usar melhor.get("nome") — pode ser fornecedor diferente do item atual
    fornecedor_eco = forn_extraido or ""
    itens_po      = po_data.get("itens") or []

    if not numero_cot:
        return {"po": numero_po, "status": "ERRO",
                "mensagem": "Número de cotação não encontrado na análise"}

    try:
        # ── A. Navegar para o histórico ─────────────────────────────────
        await page.goto(f"{ECO_URL}/requisition/history")

        # ── B. Aguardar tabela carregar ──────────────────────────────────
        search_box = page.locator("kendo-textbox input").first
        await search_box.wait_for(state="visible", timeout=TIMEOUT)
        await page.wait_for_selector("kendo-grid-list table tbody tr", timeout=TIMEOUT)
        await page.wait_for_timeout(1500)

        # ── C. Preencher campo de busca ──────────────────────────────────
        # Garante o formato XXXX.XXXXXX — insere ponto na posição 4 SOMENTE
        # para números de exatamente 10 dígitos (ex: "2025039982" → "2025.039982").
        # Outros formatos (ex: ECO REQ com 7 dígitos) são usados como estão.
        if "." not in numero_cot and len(numero_cot) == 10 and numero_cot.isdigit():
            numero_busca = numero_cot[:4] + "." + numero_cot[4:]
        else:
            numero_busca = numero_cot

        await search_box.click()
        await search_box.press("Control+a")
        await search_box.press_sequentially(numero_busca, delay=50)

        try:
            await page.wait_for_selector(
                f"kendo-grid-list table tbody tr td:has-text('{numero_busca}')",
                timeout=TIMEOUT,
            )
        except PWTimeout:
            return {
                "po": numero_po, "status": "ERRO",
                "mensagem": f"Nenhum resultado encontrado para '{numero_busca}' no histórico do ECO"
            }

        # ── D. Ler resultados ────────────────────────────────────────────
        result_cells = page.locator("td[role='gridcell'][aria-colindex='3']")
        await result_cells.first.wait_for(state="visible", timeout=TIMEOUT)
        count = await result_cells.count()

        # ── E. Selecionar a requisição ───────────────────────────────────
        if count > 1:
            # Lê os textos de todas as linhas para exibir ao usuário
            linhas_txt = []
            for idx in range(min(count, 10)):
                try:
                    txt = (await result_cells.nth(idx).inner_text()).strip()
                    linhas_txt.append(txt)
                except Exception:
                    linhas_txt.append(f"(linha {idx+1} — erro ao ler)")

            # ── CONFIRMAÇÃO 1: usuário escolhe qual requisição abrir ──────
            idx_escolhido = await _escolher_async(
                escolher,
                f"PO {numero_po} — Selecionar requisição ({count} encontrada(s))",
                linhas_txt,
            )
            if idx_escolhido < 0:
                return {"po": numero_po, "status": "CANCELADO", "mensagem": "Cancelado (seleção de requisição)"}

            row_cells = result_cells.nth(idx_escolhido)
            row_el = row_cells.locator("xpath=..")
            btn_row = row_el.locator("button")
            abriu = False
            if await btn_row.count() > 0:
                await btn_row.first.click()
                abriu = True
            if not abriu:
                await page.locator("td[role='gridcell'][aria-colindex='8'] button").first.click()
        else:
            await page.locator("td[role='gridcell'][aria-colindex='8'] button").first.click()

        # ── F. Aguardar botões Order ─────────────────────────────────────
        order_btns = page.locator("button.order.grn.card-1")
        await order_btns.first.wait_for(state="visible", timeout=TIMEOUT)
        await page.wait_for_timeout(1500)

        total_btns = await order_btns.count()

        itens_criados = 0
        po_gerado = numero_po  # valor padrão; atualizado pelo popup de confirmação

        for k in range(total_btns):
            # Os botões NÃO desaparecem do DOM após serem processados — apenas mudam
            # de texto ("Order" → "Order (1)"). Por isso usamos nth(k) para acessar
            # o botão correto de cada linha, independente do estado dos anteriores.
            btn = order_btns.nth(k)
            texto = (await btn.inner_text()).strip()
            ativo = await btn.is_enabled()

            if texto != "Order" or not ativo:
                # Botão já processado ("Order (1)") ou desabilitado — pula silenciosamente
                continue

            # Variáveis de estado do item — inicializadas aqui para evitar
            # UnboundLocalError se algum passo intermediário for pulado.
            vendor_input = None

            # ── G. Clicar em Order ───────────────────────────────────────
            await btn.scroll_into_view_if_needed()
            await page.wait_for_timeout(400)
            await btn.click()
            await page.wait_for_timeout(800)

            # ── H. Popup de confirmação (opcional) ───────────────────────
            try:
                confirm_popup = page.locator("kendo-popup button").first
                await confirm_popup.wait_for(state="visible", timeout=2000)
                await confirm_popup.click()
                await page.wait_for_timeout(500)
            except PWTimeout:
                pass  # sem popup — normal em alguns casos

            await page.wait_for_timeout(800)

            # ── H2. Aguardar formulário pronto e identificar item ──────────
            # Espera o campo Description ter valor — sinal de que o Angular
            # terminou de carregar o formulário e conectou os bindings.
            fornecedor_eco_item = fornecedor_eco  # fallback = vendor geral da PO
            desc_input = page.locator("input[formcontrolname='description']")
            try:
                await desc_input.wait_for(state="visible", timeout=TIMEOUT)
                # Polling: aguarda até o Angular preencher o campo (max ~3s)
                for _ in range(10):
                    desc_valor = (await desc_input.input_value()).strip()
                    if desc_valor:
                        break
                    await page.wait_for_timeout(300)
                else:
                    desc_valor = ""

                if desc_valor:
                    desc_norm = desc_valor.lower()
                    for it in itens_po:
                        desc_it = (it.get("descricao") or "").lower()
                        pn_it   = (it.get("pn_fornecedor") or it.get("pn") or "").lower()
                        if (desc_it and desc_it[:30] in desc_norm) or (pn_it and pn_it in desc_norm):
                            forn_it = (it.get("fornecedor_item") or "").strip()
                            if forn_it:
                                fornecedor_eco_item = forn_it
                            break
            except Exception:
                pass  # usa fornecedor_eco_item = fornecedor_eco

            # ── I. Preencher fornecedor ──────────────────────────────────
            # Seletor direto confirmado via DevTools (Angular Material autocomplete)
            vendor_input = page.locator(
                "input.mat-autocomplete-trigger[role='combobox']"
            ).first
            await vendor_input.wait_for(state="visible", timeout=TIMEOUT)
            termo_busca = _termo_busca_vendor(fornecedor_eco_item)
            # Simula ação humana: clique no campo, foca, e começa a digitar
            await vendor_input.click()
            await vendor_input.focus()
            await vendor_input.fill("")  # limpa qualquer valor anterior
            await page.keyboard.type(termo_busca, delay=30)

            # ── J. Aguardar e selecionar autocomplete ────────────────────
            chave_forn = _norm_vendor(fornecedor_eco_item)
            try:
                await page.wait_for_selector(
                    "mat-option[role='option']", timeout=SHORT_WAIT
                )
                opts = page.locator("mat-option[role='option']")
                n_opts = await opts.count()
                if n_opts > 0:
                    # Lê os textos de todas as opções
                    opcoes_txt = []
                    for oi in range(n_opts):
                        try:
                            opcoes_txt.append((await opts.nth(oi).inner_text()).strip())
                        except Exception:
                            opcoes_txt.append(f"(opção {oi+1})")

                    opcao_salva = vendor_map.get(chave_forn)
                    idx_escolhido = -1

                    if opcao_salva and opcao_salva in opcoes_txt:
                        # Já temos a escolha salva — usa diretamente
                        idx_escolhido = opcoes_txt.index(opcao_salva)
                    elif n_opts == 1:
                        # Apenas uma opção — seleciona automaticamente
                        idx_escolhido = 0
                    else:
                        # ── CONFIRMAÇÃO 2: múltiplas opções — usuário escolhe ─
                        resposta = await _escolher_async(
                            escolher,
                            f"PO {numero_po} — Selecionar fornecedor '{fornecedor_eco_item}'",
                            opcoes_txt,
                        )
                        if isinstance(resposta, str):
                            # Texto livre — limpa o campo e digita o nome fornecido
                            await vendor_input.click()
                            await vendor_input.press("Control+a")
                            await vendor_input.press_sequentially(resposta, delay=20)
                            await page.wait_for_timeout(1000)
                            # Tenta selecionar a primeira opção do autocomplete
                            try:
                                new_opt = page.locator("mat-option[role='option']").first
                                await new_opt.wait_for(state="visible", timeout=SHORT_WAIT)
                                await new_opt.click()
                            except PWTimeout:
                                pass  # sem sugestões — mantém o texto digitado
                            vendor_map[chave_forn] = resposta
                            _salvar_vendor_map(vendor_map)
                            idx_escolhido = -2  # sinaliza que já foi tratado
                        elif isinstance(resposta, int) and resposta >= 0:
                            idx_escolhido = resposta
                            vendor_map[chave_forn] = opcoes_txt[idx_escolhido]
                            _salvar_vendor_map(vendor_map)

                    if idx_escolhido >= 0:
                        await opts.nth(idx_escolhido).click()
                        await page.wait_for_timeout(500)
                else:
                    await _ok(confirmar,
                              f"PO {numero_po} — Fornecedor não encontrado",
                              "Nenhuma sugestão apareceu no autocomplete.\n"
                              "Verifique se o nome está cadastrado no ECO.\n\n"
                              "Clique OK para continuar sem selecionar.")
            except PWTimeout:
                await _ok(confirmar,
                          f"PO {numero_po} — Fornecedor não encontrado",
                          f"Timeout aguardando sugestões para '{fornecedor_eco_item}'.\n"
                          "O nome pode estar incorreto ou não cadastrado.\n\n"
                          "Clique OK para continuar sem selecionar.")

            # ── K. GL Code ───────────────────────────────────────────────
            # Pequena pausa após vendor para o Angular processar o formulário
            await page.wait_for_timeout(800)
            gl_sel = page.locator(
                "mat-select[role='combobox'][formcontrolname='itemCodeId']"
            )
            try:
                await gl_sel.wait_for(state="visible", timeout=3000)
                gl_txt = (await gl_sel.inner_text()).strip()
                if not gl_txt:
                    gl_code = vessels.get(centro_custo.upper()) if centro_custo else None
                    if gl_code:
                        await gl_sel.click()
                        await page.wait_for_timeout(500)
                        await page.keyboard.type(str(gl_code))
                        await page.wait_for_timeout(500)
                        await page.keyboard.press("Enter")
                        await page.wait_for_timeout(500)
            except PWTimeout:
                pass

            # ── L. Preço ─────────────────────────────────────────────────
            preco_preenchido = ""
            item_casado = None  # inicializa aqui para uso em L2 (Qty)
            price_input = page.locator("input[formcontrolname='price']")
            try:
                await price_input.wait_for(state="visible", timeout=3000)
                current = await price_input.input_value()
                if current:
                    preco_preenchido = current
                else:
                    desc_input = page.locator("input[formcontrolname='description']")
                    desc_eco = (await desc_input.input_value()).lower()
                    preco_encontrado = None
                    for item in itens_po:
                        desc_po = (item.get("descricao") or "").lower()
                        pn_po   = (item.get("pn") or "").lower()
                        if (desc_eco[:20] in desc_po or desc_po[:20] in desc_eco or
                                (pn_po and pn_po in desc_eco)):
                            preco_encontrado = item.get("preco_unitario")
                            item_casado = item
                            break
                    if preco_encontrado is not None:
                        await price_input.fill(str(preco_encontrado))
                        preco_preenchido = str(preco_encontrado)
            except PWTimeout:
                pass

            # ── L2. Quantidade ───────────────────────────────────────────
            qtd_preenchida = ""
            # Tenta primeiro por label (mais robusto com Angular)
            qty_input = None
            try:
                by_label = page.get_by_label("Qty", exact=False)
                await by_label.first.wait_for(state="visible", timeout=2000)
                qty_input = by_label.first
            except PWTimeout:
                pass

            # Fallback: tenta formcontrolnames conhecidos
            if qty_input is None:
                for _sel in ["input[formcontrolname='qty']",
                             "input[formcontrolname='quantity']",
                             "input[formcontrolname='qtyOrdered']",
                             "input[formcontrolname='requestedQty']",
                             "input[formcontrolname='orderedQty']"]:
                    try:
                        candidate = page.locator(_sel)
                        await candidate.wait_for(state="visible", timeout=800)
                        qty_input = candidate
                        break
                    except PWTimeout:
                        continue

            if qty_input:
                current_qty = await qty_input.input_value()
                qtd_esperada = None
                if item_casado:
                    q = item_casado.get("quantidade")
                    qtd_esperada = str(int(q)) if q is not None else None

                if qtd_esperada and current_qty != qtd_esperada:
                    await qty_input.triple_click()
                    await qty_input.fill(qtd_esperada)
                    qtd_preenchida = qtd_esperada
                else:
                    qtd_preenchida = current_qty
            # Se nenhum seletor encontrou o campo, continua sem alterar (não bloqueia)

            # ── L2b. UOM (Unidade de Medida) ───────────────────────────
            uom_extraida = ""
            if item_casado:
                uom_extraida = (item_casado.get("uom") or "").strip().lower()
            uom_eco = UOM_MAP.get(uom_extraida, "")

            if uom_eco:
                try:
                    uom_dd = page.locator(
                        "kendo-dropdownlist[formcontrolname='unitOfMeasure']"
                    ).first
                    await uom_dd.wait_for(state="visible", timeout=3000)
                    # Verifica valor atual — só altera se diferente
                    uom_atual = (await uom_dd.inner_text()).strip().lower()
                    if uom_atual != uom_eco:
                        await uom_dd.click()
                        await page.wait_for_timeout(400)
                        # Kendo popup: li items na lista
                        uom_opt = page.locator("kendo-popup li, ul.k-list li").filter(
                            has_text=re.compile(f"^{re.escape(uom_eco)}$", re.IGNORECASE)
                        ).first
                        await uom_opt.wait_for(state="visible", timeout=3000)
                        await uom_opt.click()
                        await page.wait_for_timeout(300)
                except Exception:
                    pass  # Mantém UOM padrão se não conseguir selecionar

            # ── L3. Ship VIA ─────────────────────────────────────────────
            # Prioridade: valor direto da planilha (col V) > mapeamento tipo_freight
            tipo_freight   = (melhor.get("tipo_freight") or "").strip()
            ship_via_alvo  = (melhor.get("ship_via_direto") or "").strip() or SHIP_VIA_MAP.get(tipo_freight.lower())

            if ship_via_alvo:
                _sv_filled = False
                _pattern = re.compile(re.escape(ship_via_alvo), re.IGNORECASE)

                # Campo confirmado via DevTools: mat-select[formcontrolname="shipViaId"]
                try:
                    sv = page.locator("mat-select[formcontrolname='shipViaId']").first
                    await sv.wait_for(state="visible", timeout=3000)
                    await sv.click()
                    await page.wait_for_timeout(600)
                    opt = page.locator("mat-option").filter(has_text=_pattern).first
                    await opt.wait_for(state="visible", timeout=3000)
                    await opt.click()
                    _sv_filled = True
                except Exception:
                    pass

                if not _sv_filled:
                    await _ok(confirmar,
                              f"PO {numero_po} — Ship VIA não preenchido",
                              f"Não foi possível selecionar '{ship_via_alvo}'.\n\n"
                              "Preencha manualmente e clique OK para continuar.")
                await page.wait_for_timeout(400)
            else:
                await _ok(confirmar,
                          f"PO {numero_po} — Ship VIA — preencher manualmente",
                          f"Tipo de frete '{tipo_freight or '(não informado)'}' não tem mapeamento "
                          "automático para Ship VIA.\n\n"
                          "Opções disponíveis no ECO:\n"
                          "  • ECO UPS ACCT# 707185\n"
                          "  • Free delivery\n"
                          "  • Runner Pick up\n"
                          "  • Supplier Ship\n\n"
                          "Preencha manualmente no navegador e clique OK para continuar.")

            # ── M. Confirmar e submeter ──────────────────────────────────
            vendor_atual = fornecedor_eco_item
            try:
                if vendor_input is not None:
                    vendor_atual = await vendor_input.input_value()
            except Exception:
                pass  # mantém fornecedor_eco como fallback

            if not await _ok(confirmar,
                             f"PO {numero_po} — M: SUBMETER item {k+1}/{total_btns}",
                             f"RESUMO DO ITEM A SUBMETER:\n\n"
                             f"  Fornecedor : {vendor_atual or '(vazio)'}\n"
                             f"  Quantidade : {qtd_preenchida or '(não alterada)'}\n"
                             f"  Preço      : {preco_preenchido or '(vazio / já preenchido)'}\n"
                             f"  Centro     : {centro_custo or '(não informado)'}\n\n"
                             "Clique SIM para salvar ou NÃO para fechar sem salvar."):
                # Fecha o modal sem salvar (tecla Escape ou botão ×)
                try:
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(1000)
                except Exception:
                    pass
                continue

            submit = page.get_by_role("button", name="Save")
            await submit.wait_for(state="visible", timeout=TIMEOUT)
            await submit.click()

            # Aguarda popup de confirmação "purchase order have been generated successfully"
            # e clica em Close para fechar
            try:
                close_btn = page.get_by_role("button", name="Close")
                await close_btn.wait_for(state="visible", timeout=10000)
                # Tenta capturar o número da PO do popup
                try:
                    popup_txt = await page.locator("text=/PO Number:/").inner_text(timeout=2000)
                    po_gerado_popup = popup_txt.split("PO Number:")[-1].strip().split()[0]
                    if po_gerado_popup:
                        po_gerado = po_gerado_popup
                except Exception:
                    pass
                await close_btn.click()
                await page.wait_for_timeout(1000)
            except PWTimeout:
                # Popup não apareceu — continua normalmente
                await page.wait_for_timeout(1500)

            itens_criados += 1

        # ── N. Capturar número da PO gerada ─────────────────────────────
        try:
            po_span = page.locator("app-requisition-header-toggle h1 span").first
            await po_span.wait_for(state="visible", timeout=5000)
            po_gerado = (await po_span.inner_text()).strip()
        except PWTimeout:
            po_gerado = numero_po

        return {
            "po": po_gerado,
            "status": "OK",
            "mensagem": f"{itens_criados} item(ns) processado(s)"
        }

    except Exception as exc:
        return {
            "po": numero_po,
            "status": "ERRO",
            "mensagem": str(exc)[:200]
        }


# ─────────────────────────────────────────────────────────────────────
# Runner principal
# ─────────────────────────────────────────────────────────────────────

async def _run(usuario: str, senha: str, lote: list, callback, confirmar, escolher) -> list:
    vessels    = _carregar_vessels()
    vendor_map = _carregar_vendor_map()
    resultados = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            slow_mo=100,
        )
        context = await browser.new_context()
        page    = await context.new_page()
        page.set_default_timeout(TIMEOUT)

        callback("Fazendo login no ECO Requisition...")
        await _login(page, usuario, senha, confirmar)
        callback("Login OK. Iniciando criação de POs...")

        for i, par in enumerate(lote):
            numero_po = (par["analise"].get("po") or {}).get("numero_po") or f"Par {i+1}"
            callback(f"[{i+1}/{len(lote)}] Processando PO {numero_po}...")
            resultado = await _criar_po_par(page, par, vessels, confirmar, escolher, vendor_map)
            resultados.append(resultado)

        await browser.close()

    return resultados


def criar_pos(usuario: str, senha: str, lote: list, callback,
              confirmar=None, escolher=None) -> list:
    """
    Ponto de entrada síncrono chamado pelo tkinter em thread separada.

    confirmar: callable(titulo, mensagem) → bool
        Confirmações passo a passo durante testes. None = autônomo.

    escolher: callable(titulo, opcoes: list[str]) → int
        Exibe lista numerada e retorna índice escolhido (0-based), -1 = cancelar.
        None = autônomo (seleciona primeira opção ou usa vendor_map).
    """
    return asyncio.run(_run(usuario, senha, lote, callback, confirmar, escolher))
