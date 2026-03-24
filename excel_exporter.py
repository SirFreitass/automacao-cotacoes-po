"""
excel_exporter.py
-----------------
Gera um único arquivo Excel para todas as análises do lote.
Estrutura:
  Aba "Índice"          : tabela com todas as análises do lote
  Aba "N-Resumo"        : comparativo de cotações (par N)
  Aba "N-Alertas"       : divergências PO (par N)
  Aba "N-Dados"         : dados estruturados para VBA (par N)
"""

import logging
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import ROBO_PLANILHA, SHIP_VIA_MAP
from utils import norm_pn as _norm_pn, norm_vendor as _norm_vendor, numero_cotacao as _numero_cotacao, quotation_code as _quotation_code, normalizar_freight_robo as _normalizar_freight_robo

logger = logging.getLogger(__name__)


# --- Lookup de fornecedores via Tabela Forn do Req-o-matic ---
# Carregado uma única vez na inicialização do módulo para evitar
# reabrir o arquivo a cada exportação.
_TABELA_FORN: dict = {}   # NomeExtraido.lower() → NomeSistema

def _carregar_tabela_forn():
    """Lê a aba Tabela Forn do Req-o-matic e monta o dicionário de lookup."""
    global _TABELA_FORN
    if _TABELA_FORN:
        return  # já carregado
    try:
        wb = load_workbook(ROBO_PLANILHA, read_only=True, data_only=True)
        ws = wb["Tabela Forn"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome_extraido = row[0]   # col A — NomeExtraido
            nome_sistema = row[2]    # col C — NomeSistema
            if nome_extraido and nome_sistema:
                _TABELA_FORN[str(nome_extraido).lower().strip()] = str(nome_sistema).strip()
        wb.close()
    except Exception:
        logger.warning("Não foi possível carregar Tabela Forn do Req-o-matic: %s", ROBO_PLANILHA)


def _carregar_vendor_map_json() -> dict:
    """Carrega o vendor_map.json para busca nas observações."""
    try:
        import json as _json
        caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor_map.json")
        with open(caminho, "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


def _buscar_fornecedor_nas_obs(observacoes: str) -> str:
    """
    Procura nomes de fornecedores conhecidos dentro do texto das observações da PO.
    Busca em duas fontes: Tabela Forn (Req-o-matic) e vendor_map.json.
    Retorna o nome encontrado (prefere o mais longo/específico).
    """
    if not observacoes:
        return ""
    _carregar_tabela_forn()
    obs_norm = _norm_vendor(observacoes)
    melhor_nome = ""
    melhor_tam = 0

    # 1. Busca na Tabela Forn do Req-o-matic
    for k, v in _TABELA_FORN.items():
        k_norm = _norm_vendor(k)
        if k_norm and len(k_norm) >= 4 and k_norm in obs_norm:
            if len(k_norm) > melhor_tam:
                melhor_tam = len(k_norm)
                melhor_nome = v

    # 2. Busca no vendor_map.json (chave = nome curto, valor = nome ECO completo)
    for k, v in _carregar_vendor_map_json().items():
        k_norm = _norm_vendor(k)
        if k_norm and len(k_norm) >= 4 and k_norm in obs_norm:
            if len(k_norm) > melhor_tam:
                melhor_tam = len(k_norm)
                melhor_nome = v

    return melhor_nome


def _palavras_sig(s: str):
    """Palavras com 3+ caracteres de uma string (para matching de fornecedor)."""
    return {w for w in re.split(r'\W+', s.lower()) if len(w) >= 3}


def _lookup_fornecedor_eco(nome_extraido: str) -> str:
    """
    Dado o nome extraído dos comentários da PO, retorna o nome do
    fornecedor no sistema ECO (Tabela Forn → NomeSistema).
    Busca em 3 níveis:
      1. Exata normalizada (sem separadores): "kmar" == "k-mar" == "K MAR"
      2. Substring normalizada: "kmar" encontra "kmarsupply"
      3. Palavras significativas com score ≥ 50%
    """
    if not nome_extraido:
        return ""
    _carregar_tabela_forn()
    norm_entrada = _norm_vendor(nome_extraido)
    # 1. Busca exata normalizada
    for k, v in _TABELA_FORN.items():
        if _norm_vendor(k) == norm_entrada:
            return v
    # 2. Busca por substring normalizada (entrada contém chave ou vice-versa)
    for k, v in _TABELA_FORN.items():
        norm_k = _norm_vendor(k)
        if norm_k and norm_entrada and (norm_k in norm_entrada or norm_entrada in norm_k):
            return v
    # 3. Busca por palavras significativas com score ≥ 50%
    palavras_entrada = _palavras_sig(nome_extraido.lower().strip())
    if not palavras_entrada:
        return ""
    melhor_v, melhor_score = "", 0.0
    for k, v in _TABELA_FORN.items():
        palavras_k = _palavras_sig(k)
        if not palavras_k:
            continue
        comuns = palavras_entrada & palavras_k
        if not comuns:
            continue
        score = len(comuns) / max(len(palavras_entrada), len(palavras_k))
        if score > melhor_score:
            melhor_score = score
            melhor_v = v
    return melhor_v if melhor_score >= 0.5 else ""


# --- Paleta de cores ---
COR_CABECALHO   = "1F3864"
COR_MELHOR      = "C6EFCE"
COR_MELHOR_FONT = "276221"
COR_ALERTA      = "FFCCCC"
COR_ALERTA_FONT = "9C0006"
COR_AVISO       = "FFEB9C"
COR_AVISO_FONT  = "9C6500"
COR_LINHA_PAR   = "F2F2F2"
COR_BRANCO      = "FFFFFF"
COR_INDICE      = "2E539E"

_BORDA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _cabecalho(ws, linha, colunas):
    for col, titulo in enumerate(colunas, 1):
        c = ws.cell(row=linha, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA


def _celula(ws, linha, col, valor, cor_fundo=None, negrito=False, cor_fonte=None,
            formato=None, alinhamento="left"):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(bold=negrito, color=cor_fonte or "000000", size=10)
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=True)
    c.border = _BORDA
    if formato:
        c.number_format = formato
    return c


def _auto_largura(ws, col, min_w=10, max_w=50):
    col_letter = get_column_letter(col)
    maior = min_w
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if c.value:
                maior = min(max(maior, len(str(c.value)) + 2), max_w)
    ws.column_dimensions[col_letter].width = maior


# =====================================================================
# ABA ÍNDICE
# =====================================================================

def _aba_indice(wb, lote):
    ws = wb.create_sheet("Índice", 0)

    # Título
    n_cols = 8
    ws.merge_cells(f"A1:H1")
    t = ws["A1"]
    t.value = f"LOTE DE ANÁLISES — {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   {len(lote)} par(es) analisado(s)"
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_INDICE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    _cabecalho(ws, 2, ["#", "Nº ECO REQ", "Nº PO", "Nº Cotação", "Melhor Fornecedor",
                        "Melhor Preço (USD)", "Alertas PO", "Aba"])
    ws.row_dimensions[2].height = 30

    for i, entrada in enumerate(lote):
        analise = entrada["analise"]
        req_numero = entrada.get("req_numero")
        n = i + 1
        linha = n + 2

        po = analise.get("po", {})
        melhor = analise.get("melhor_preco") or {}
        n_alertas = len(analise.get("alertas_po", []))

        # Nº REQ: tenta do documento, depois do nome do arquivo
        eco_req = (po.get("numero_eco_req")
                   or _req_de_cotacao(analise)
                   or req_numero
                   or "—")
        numero_po = po.get("numero_po") or "—"
        numero_cot = _quotation_code(analise) or "—"

        cor = COR_LINHA_PAR if i % 2 == 0 else COR_BRANCO
        cor_alerta = COR_ALERTA if n_alertas > 0 else COR_MELHOR

        dados = [n, eco_req, numero_po, numero_cot,
                 melhor.get("nome", "—"),
                 melhor.get("preco_total"),
                 f"{n_alertas} alerta(s)",
                 f"{n}-Resumo"]

        formatos = [None, None, None, None, None, '"$"#,##0.00', None, None]
        alins = ["center", "center", "center", "center", "left", "right", "center", "center"]

        for col, (val, fmt, aln) in enumerate(zip(dados, formatos, alins), 1):
            cf = cor_alerta if col == 7 else cor
            _celula(ws, linha, col, val, cf, formato=fmt, alinhamento=aln)
        ws.row_dimensions[linha].height = 18

    for col in range(1, n_cols + 1):
        _auto_largura(ws, col)
    ws.column_dimensions["E"].width = 30


def _req_de_cotacao(analise):
    """Pega o ECO REQ do primeiro fornecedor que tiver."""
    for forn in analise.get("ranking_preco", []):
        nr = forn.get("numero_eco_req")
        if nr:
            return nr
    return None


# =====================================================================
# ABA RESUMO ANÁLISE (por par)
# =====================================================================

def _aba_resumo(wb, analise, prefixo):
    ws = wb.create_sheet(f"{prefixo}-Resumo")
    ws.freeze_panes = "A3"

    po = analise.get("po", {})
    fornecedores = analise.get("resumo_fornecedores", [])
    melhor_preco_nome = (analise.get("melhor_preco") or {}).get("nome", "")
    melhor_prazo_nome = (analise.get("melhor_prazo") or {}).get("nome", "")

    eco_req = po.get("numero_eco_req") or _req_de_cotacao(analise) or "—"
    numero_po = po.get("numero_po") or "—"
    numero_cot = _quotation_code(analise) or "—"

    titulo = (f"ANÁLISE DE COTAÇÕES  |  REQ: {eco_req}  |  PO: {numero_po}  "
              f"|  Cotação: {numero_cot}  |  {datetime.now().strftime('%d/%m/%Y')}")
    n_cols = 14
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = titulo
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    colunas = ["Rank Preço", "Rank Prazo", "Fornecedor", "Preço Total (USD)",
               "Custo Freight (USD)", "Tipo Freight", "Prazo Entrega", "Prazo (dias)",
               "Forma Pagamento", "Item Substituto?", "Nº Cotação", "Validade Cotação",
               "Melhor Preço?", "Melhor Prazo?"]
    _cabecalho(ws, 2, colunas)
    ws.row_dimensions[2].height = 32

    for i, forn in enumerate(fornecedores):
        linha = i + 3
        cor = COR_LINHA_PAR if i % 2 == 0 else COR_BRANCO
        nome = forn.get("nome", "")
        eh_melhor_preco = nome == melhor_preco_nome
        eh_melhor_prazo = nome == melhor_prazo_nome
        tem_sub = forn.get("tem_item_substituto") == "SIM"
        cor_linha = COR_MELHOR if eh_melhor_preco else cor

        dados = [forn.get("posicao_preco"), forn.get("posicao_prazo"), nome,
                 forn.get("preco_total"), forn.get("custo_freight"), forn.get("tipo_freight"),
                 forn.get("prazo_entrega"), forn.get("prazo_entrega_dias"),
                 forn.get("forma_pagamento"), forn.get("tem_item_substituto"),
                 forn.get("numero_cotacao"), forn.get("validade_cotacao"),
                 "✓ MELHOR" if eh_melhor_preco else "",
                 "✓ MELHOR" if eh_melhor_prazo else ""]
        formatos = [None, None, None, '"$"#,##0.00', '"$"#,##0.00', None, None,
                    None, None, None, None, None, None, None]
        alins = ["center", "center", "left", "right", "right", "center", "center",
                 "center", "center", "center", "center", "center", "center", "center"]

        validade_vencida = forn.get("validade_vencida")
        validade_dias = forn.get("validade_dias_restantes")

        for col, (val, fmt, aln) in enumerate(zip(dados, formatos, alins), 1):
            cf = cor_linha
            cf_font = None
            neg = False
            if col == 10 and tem_sub:
                cf = COR_AVISO; cf_font = COR_AVISO_FONT; neg = True
            if col == 12:
                if validade_vencida:
                    cf = COR_ALERTA; cf_font = COR_ALERTA_FONT; neg = True
                elif validade_dias is not None and validade_dias <= 7:
                    cf = COR_AVISO; cf_font = COR_AVISO_FONT; neg = True
                elif validade_dias is not None:
                    cf = COR_MELHOR; cf_font = COR_MELHOR_FONT; neg = True
            if col in (13, 14) and val:
                cf = COR_MELHOR; cf_font = COR_MELHOR_FONT; neg = True
            _celula(ws, linha, col, val, cf, neg, cf_font, fmt, aln)
        ws.row_dimensions[linha].height = 18

    # Alertas de itens substitutos
    alertas_itens = analise.get("alertas_itens", [])
    if alertas_itens:
        lr = len(fornecedores) + 4
        ws.merge_cells(f"A{lr}:{get_column_letter(n_cols)}{lr}")
        c = ws.cell(row=lr, column=1, value="⚠ ITENS SIMILARES / SUBSTITUTOS")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="FF9900")
        c.alignment = Alignment(horizontal="center")
        lr += 1
        _cabecalho(ws, lr, ["Fornecedor", "Part Number", "Descrição", "Observação"])
        ws.merge_cells(f"D{lr}:{get_column_letter(n_cols)}{lr}")
        lr += 1
        for alerta in alertas_itens:
            _celula(ws, lr, 1, alerta.get("fornecedor"), COR_AVISO)
            _celula(ws, lr, 2, alerta.get("pn"), COR_AVISO)
            _celula(ws, lr, 3, alerta.get("descricao"), COR_AVISO)
            c = ws.cell(row=lr, column=4, value=alerta.get("observacao"))
            c.fill = PatternFill("solid", fgColor=COR_AVISO)
            c.font = Font(color=COR_AVISO_FONT)
            ws.merge_cells(f"D{lr}:{get_column_letter(n_cols)}{lr}")
            lr += 1

    for col in range(1, n_cols + 1):
        _auto_largura(ws, col)


# =====================================================================
# ABA ALERTAS PO (por par)
# =====================================================================

def _aba_alertas(wb, analise, prefixo):
    ws = wb.create_sheet(f"{prefixo}-Alertas")
    ws.freeze_panes = "A3"

    po = analise.get("po", {})
    alertas = analise.get("alertas_po", [])

    eco_req = po.get("numero_eco_req") or _req_de_cotacao(analise) or "—"
    numero_po = po.get("numero_po") or "—"

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"VALIDAÇÃO DA PO  |  PO: {numero_po}  |  REQ: {eco_req}  |  {datetime.now().strftime('%d/%m/%Y')}"
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    _cabecalho(ws, 2, ["Tipo", "Severidade", "Mensagem", "", ""])
    ws.merge_cells("C2:E2")
    ws.row_dimensions[2].height = 28

    if not alertas:
        ws.merge_cells("A3:E3")
        c = ws.cell(row=3, column=1, value="✓ Nenhum alerta. PO está em conformidade com a cotação.")
        c.font = Font(bold=True, color=COR_MELHOR_FONT, size=11)
        c.fill = PatternFill("solid", fgColor=COR_MELHOR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22
    else:
        for i, alerta in enumerate(alertas):
            linha = i + 3
            sev = alerta.get("severidade", "INFO")
            cor = COR_ALERTA if sev == "ALERTA" else (COR_AVISO if sev == "AVISO" else COR_LINHA_PAR)
            cf = COR_ALERTA_FONT if sev == "ALERTA" else (COR_AVISO_FONT if sev == "AVISO" else None)
            _celula(ws, linha, 1, alerta.get("tipo"), cor, True, cf, alinhamento="center")
            _celula(ws, linha, 2, sev, cor, True, cf, alinhamento="center")
            c = ws.cell(row=linha, column=3, value=alerta.get("mensagem"))
            c.font = Font(color=cf or "000000", size=10)
            c.fill = PatternFill("solid", fgColor=cor)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = _BORDA
            ws.merge_cells(f"C{linha}:E{linha}")
            ws.row_dimensions[linha].height = 30

    # Info da PO
    lr = len(alertas) + 5 if alertas else 5
    ws.merge_cells(f"A{lr}:E{lr}")
    c = ws.cell(row=lr, column=1, value="DADOS DA PO")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    c.alignment = Alignment(horizontal="center")
    lr += 1
    for campo, valor in [
        ("Nº ECO REQ", eco_req),
        ("Número da PO", numero_po),
        ("Fornecedor (cabeçalho)", po.get("fornecedor_selecionado")),
        ("Fornecedor escolhido (comentários)", po.get("fornecedor_escolhido_comentario")),
        ("Subtotal", po.get("subtotal")),
        ("Custo Freight", po.get("custo_freight")),
        ("Total", po.get("preco_total")),
        ("Forma de Pagamento", po.get("forma_pagamento")),
        ("Observações", po.get("observacoes")),
    ]:
        c1 = ws.cell(row=lr, column=1, value=campo)
        c1.font = Font(bold=True, size=10)
        c1.fill = PatternFill("solid", fgColor=COR_LINHA_PAR)
        c1.border = _BORDA
        c2 = ws.cell(row=lr, column=2, value=valor)
        c2.border = _BORDA
        c2.font = Font(size=10)
        ws.merge_cells(f"B{lr}:E{lr}")
        lr += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60


# =====================================================================
# ABA DADOS VBA (por par)
# =====================================================================

def _aba_dados_vba(wb, analise, prefixo):  # mantida para compatibilidade mas não chamada
    ws = wb.create_sheet(f"{prefixo}-Dados")

    po = analise.get("po", {})
    melhor = analise.get("melhor_preco") or {}

    eco_req = po.get("numero_eco_req") or _req_de_cotacao(analise) or ""
    numero_cot = _numero_cotacao(analise) or ""

    colunas = [
        "ECO_REQ", "PO_NUMERO", "PO_COTACAO", "PO_FORNECEDOR", "PO_FORNECEDOR_ESCOLHIDO",
        "PO_PRECO_TOTAL", "PO_FREIGHT", "PO_PAGAMENTO", "PO_PRAZO",
        "ITEM_LINHA", "ITEM_PN_INTERNO", "ITEM_PN_FORNECEDOR", "ITEM_DESCRICAO",
        "ITEM_QTD", "ITEM_PRECO_UNIT", "ITEM_PRECO_TOTAL",
        "COT_MELHOR_FORN", "COT_MELHOR_PRECO", "COT_MELHOR_PRAZO",
        "COT_TIPO_FREIGHT", "COT_PAGAMENTO",
        "STATUS_FREIGHT_OK", "STATUS_PN_OK",
    ]
    _cabecalho(ws, 1, colunas)
    ws.row_dimensions[1].height = 32

    alertas_po = analise.get("alertas_po", [])
    tipos_alerta = {a.get("tipo") for a in alertas_po}
    freight_ok = "OK" if "FREIGHT" not in tipos_alerta else "VERIFICAR"

    for i, item in enumerate(po.get("itens") or []):
        linha = i + 2
        pn_interno = item.get("pn") or ""
        pn_forn = item.get("pn_fornecedor") or ""
        pn_busca = (pn_forn or pn_interno).upper()
        itens_melhor = {(it.get("pn") or "").upper(): it for it in melhor.get("itens", []) if it.get("pn")}
        item_ref = itens_melhor.get(pn_busca) or next(
            (v for k, v in itens_melhor.items() if pn_busca in k or k in pn_busca), {}
        )
        pn_ok = "OK" if item_ref else "VERIFICAR"
        cor = COR_LINHA_PAR if i % 2 == 0 else COR_BRANCO

        dados = [
            eco_req, po.get("numero_po"), numero_cot,
            po.get("fornecedor_selecionado"), po.get("fornecedor_escolhido_comentario"),
            po.get("preco_total"), po.get("custo_freight"),
            po.get("forma_pagamento"), po.get("prazo_entrega"),
            i + 1, pn_interno, pn_forn, item.get("descricao"),
            item.get("quantidade"), item.get("preco_unitario"), item.get("preco_total_item"),
            melhor.get("nome"), melhor.get("preco_total"), melhor.get("prazo_entrega"),
            melhor.get("tipo_freight"), melhor.get("forma_pagamento"),
            freight_ok, pn_ok,
        ]
        formatos = [None]*9 + [None, None, None, None, None,
                    '"$"#,##0.00', '"$"#,##0.00', None, '"$"#,##0.00',
                    None, None, None, None, None]

        for col, (val, fmt) in enumerate(zip(dados, formatos), 1):
            cf = COR_ALERTA if (col in (22, 23) and val == "VERIFICAR") else cor
            _celula(ws, linha, col, val, cf, formato=fmt)
        ws.row_dimensions[linha].height = 18

    for col in range(1, len(colunas) + 1):
        _auto_largura(ws, col)


# =====================================================================
# ABA PARA ROBO — colunas idênticas à planilha pos do Req-o-matic
# =====================================================================

def _aba_para_robo(wb, analise, prefixo):
    """
    Gera aba com a mesma estrutura de colunas da aba 'pos' do Req-o-matic.
    O usuário pode copiar estas linhas diretamente para o Req-o-matic.

    Mapeamento de colunas (A→S):
      A  Quotation Code       → numero_cotacao
      B  Produto              → pn_interno (código interno da PO)
      C  Description          → descricao do item
      D  Unit Price           → preco_unitario
      E  Cost Center Desc     → centro_de_custo da seção "Cost Center Apportionment" (ex: "C-ADMIRAL")
      F  Supplier             → nome ECO via Tabela Forn (lookup por fornecedor_escolhido_comentario)
      G  Freight              → freight normalizado para vocabulário ECO
      H  Status               → (vazio — preenchido pelo VBA)
      I  OBS                  → alertas da análise (se houver)
      J  ID+QUOTE             → pn_interno + numero_cotacao (concatenação)
      K  ECO REQ              → numero_eco_req
      L  Quote+PO             → numero_po + numero_cotacao (concatenação)
      M  Coluna3              → (vazio)
      N  Coluna4              → (vazio)
      O  Observação PO        → observacoes da PO (comentários do comprador)
      P  Forn. Extraido       → fornecedor_escolhido_comentario
      Q  Forn Extraído ECO    → (vazio — VLOOKUP do Req-o-matic preenche)
      R  PO                   → numero_po
      S  Coluna5              → resumo: "PO:{po} - {cotacao} - {centro_de_custo}"
    """
    ws = wb.create_sheet(f"{prefixo}-Robo")

    po = analise.get("po", {})
    alertas_po = analise.get("alertas_po", [])

    eco_req = po.get("numero_eco_req") or _req_de_cotacao(analise) or ""
    # Fallback: se a cotação não trouxe o número, tenta o campo extraído da própria PO
    numero_cot = _quotation_code(analise)
    numero_po = po.get("numero_po") or ""
    # Centro de custo: campo extraído da seção "Cost Center Apportionment" da PO
    # Ex: "(0185) C-ADMIRAL - USD 3.500,00" → "C-ADMIRAL"
    centro_de_custo = po.get("centro_de_custo") or po.get("solicitante") or ""
    forn_extraido = po.get("fornecedor_escolhido_comentario") or ""
    observacoes = po.get("observacoes") or ""

    # Fallback em cadeia para encontrar o fornecedor:
    # 1. fornecedor_escolhido_comentario (Gemini)
    # 2. Busca de nomes conhecidos nas observações da PO
    # 3. Fornecedor do melhor preço da cotação
    if not forn_extraido:
        forn_extraido = _buscar_fornecedor_nas_obs(observacoes)
    if not forn_extraido:
        melhor = analise.get("melhor_preco") or {}
        forn_extraido = melhor.get("nome") or ""

    # Lookup do nome do fornecedor no sistema ECO via Tabela Forn do Req-o-matic
    fornecedor_eco = _lookup_fornecedor_eco(forn_extraido) or forn_extraido or ""

    # Freight normalizado para vocabulário do Req-o-matic
    referencia = analise.get("melhor_preco") or {}
    freight_robo = _normalizar_freight_robo(
        referencia.get("tipo_freight"), referencia.get("custo_freight")
    )

    # Resumo de alertas para coluna OBS
    obs_alertas = "; ".join(
        f"[{a.get('tipo')}] {a.get('mensagem', '')[:60]}"
        for a in alertas_po
    ) if alertas_po else ""

    # Cabeçalhos exatamente como na planilha pos
    colunas = [
        "Quotation Code", "Produto", "Description", "Unit Price",
        "Cost Center Desc", "Supplier", "Freight", "Status", "OBS",
        "ID+QUOTE", "ECO REQ", "Quote+PO", "Coluna3", "Coluna4",
        "Observação PO", "Forn. Extraido", "Forn Extraído ECO", "PO", "Coluna5",
    ]

    # Cabeçalho especial com cor diferente para destacar que é a aba de importação
    for col, titulo in enumerate(colunas, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="2E539E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA
    ws.row_dimensions[1].height = 32

    # Instrução na linha 2
    ws.merge_cells("A2:S2")
    inst = ws["A2"]
    inst.value = (
        "⬇  Copie as linhas abaixo e cole na aba 'pos' do Req-o-matic para registrar no robô  ⬇"
    )
    inst.font = Font(bold=True, color=COR_AVISO_FONT, size=10)
    inst.fill = PatternFill("solid", fgColor=COR_AVISO)
    inst.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    itens = po.get("itens") or []
    if not itens:
        # Sem itens: gera ao menos uma linha com dados gerais
        itens = [{}]

    for i, item in enumerate(itens):
        linha = i + 3
        pn_interno = item.get("pn") or ""
        descricao = item.get("descricao") or ""
        preco_unit = item.get("preco_unitario")

        # Vendor por item: usa fornecedor_item se disponível, senão usa o vendor geral da PO
        forn_it = (item.get("fornecedor_item") or "").strip()
        fornecedor_eco_item = _lookup_fornecedor_eco(forn_it) or forn_it or fornecedor_eco

        id_quote = f"{pn_interno}{numero_cot}"
        quote_po = f"{numero_po}{numero_cot}"
        coluna5 = f"PO:{numero_po} - {numero_cot} - {centro_de_custo}"

        dados = [
            numero_cot,           # A  Quotation Code
            pn_interno,           # B  Produto
            descricao,            # C  Description
            preco_unit,           # D  Unit Price
            centro_de_custo,      # E  Cost Center Desc  (ex: "C-ADMIRAL")
            fornecedor_eco_item,  # F  Supplier  (por item se disponível, senão PO geral)
            freight_robo,     # G  Freight
            "",               # H  Status (VBA preenche)
            obs_alertas,      # I  OBS
            id_quote,         # J  ID+QUOTE
            eco_req,          # K  ECO REQ
            quote_po,         # L  Quote+PO
            "",               # M  Coluna3
            "",               # N  Coluna4
            observacoes,      # O  Observação PO
            forn_extraido,    # P  Forn. Extraido
            "",               # Q  Forn Extraído ECO (VLOOKUP do Req-o-matic)
            numero_po,        # R  PO
            coluna5,          # S  Coluna5
        ]

        cor = COR_LINHA_PAR if i % 2 == 0 else COR_BRANCO
        cor_obs = COR_AVISO if obs_alertas else cor

        for col, val in enumerate(dados, 1):
            cf = cor_obs if col == 9 and obs_alertas else cor
            fmt = '"$"#,##0.0000' if col == 4 else None
            _celula(ws, linha, col, val, cf, formato=fmt)
        ws.row_dimensions[linha].height = 18

    # Larguras fixas para refletir a planilha pos
    larguras = {
        "A": 16, "B": 14, "C": 45, "D": 14, "E": 20, "F": 28,
        "G": 16, "H": 14, "I": 35, "J": 28, "K": 18, "L": 28,
        "M": 10, "N": 10, "O": 45, "P": 25, "Q": 25, "R": 12, "S": 40,
    }
    for col_letter, w in larguras.items():
        ws.column_dimensions[col_letter].width = w


# =====================================================================
# ABA ANÁLISE — Resumo + Alertas + Comparativo de Itens (por par)
# =====================================================================

def _aba_analise(wb, analise, prefixo):
    """Aba unificada: comparativo de fornecedores, alertas e comparativo de itens PO vs cotação."""
    ws = wb.create_sheet(f"{prefixo}-Análise")
    ws.freeze_panes = "A3"

    po             = analise.get("po", {})
    fornecedores   = analise.get("resumo_fornecedores", [])
    alertas        = analise.get("alertas_po", [])
    melhor         = analise.get("melhor_preco") or {}
    melhor_nome    = melhor.get("nome", "")
    melhor_prazo_nome = (analise.get("melhor_prazo") or {}).get("nome", "")

    eco_req    = po.get("numero_eco_req") or _req_de_cotacao(analise) or "—"
    numero_po  = po.get("numero_po") or "—"
    numero_cot = _quotation_code(analise) or "—"
    n_alertas  = len(alertas)
    cor_titulo = COR_ALERTA if n_alertas else COR_MELHOR

    # ── Título ──
    n_cols = 14
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = (f"ANÁLISE  |  REQ: {eco_req}  |  PO: {numero_po}  |  "
               f"Cotação: {numero_cot}  |  {n_alertas} alerta(s)  |  "
               f"{datetime.now().strftime('%d/%m/%Y')}")
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    lr = 2  # linha corrente

    # ── Seção 1: Comparativo de Fornecedores ──
    ws.merge_cells(f"A{lr}:{get_column_letter(n_cols)}{lr}")
    c = ws.cell(row=lr, column=1, value="COMPARATIVO DE FORNECEDORES")
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[lr].height = 18
    lr += 1

    _cabecalho(ws, lr, ["Rank Preço", "Rank Prazo", "Fornecedor", "Preço Total (USD)",
                         "Freight (USD)", "Tipo Freight", "Prazo Entrega", "Prazo (dias)",
                         "Pagamento", "Item Substituto?", "Nº Cotação", "Validade",
                         "Melhor Preço?", "Melhor Prazo?"])
    ws.row_dimensions[lr].height = 32
    lr += 1

    for i, forn in enumerate(fornecedores):
        cor = COR_MELHOR if forn.get("nome") == melhor_nome else (COR_LINHA_PAR if i % 2 == 0 else COR_BRANCO)
        tem_sub = forn.get("tem_item_substituto") == "SIM"
        dados = [forn.get("posicao_preco"), forn.get("posicao_prazo"), forn.get("nome"),
                 forn.get("preco_total"), forn.get("custo_freight"), forn.get("tipo_freight"),
                 forn.get("prazo_entrega"), forn.get("prazo_entrega_dias"),
                 forn.get("forma_pagamento"), forn.get("tem_item_substituto"),
                 forn.get("numero_cotacao"), forn.get("validade_cotacao"),
                 "✓ MELHOR" if forn.get("nome") == melhor_nome else "",
                 "✓ MELHOR" if forn.get("nome") == melhor_prazo_nome else ""]
        fmts = [None,None,None,'"$"#,##0.00','"$"#,##0.00',None,None,None,None,None,None,None,None,None]
        alns = ["center","center","left","right","right","center","center",
                "center","center","center","center","center","center","center"]
        for col, (val, fmt, aln) in enumerate(zip(dados, fmts, alns), 1):
            cf = COR_AVISO if col == 10 and tem_sub else cor
            ff = COR_AVISO_FONT if col == 10 and tem_sub else (COR_MELHOR_FONT if col in (13,14) and val else None)
            _celula(ws, lr, col, val, cf, negrito=bool(ff), cor_fonte=ff, formato=fmt, alinhamento=aln)
        ws.row_dimensions[lr].height = 18
        lr += 1

    lr += 1  # espaço

    # ── Seção 2: Alertas da PO ──
    ws.merge_cells(f"A{lr}:{get_column_letter(n_cols)}{lr}")
    cab_alerta = ws.cell(row=lr, column=1,
                         value="ALERTAS DA PO" if alertas else "✓ PO EM CONFORMIDADE — Nenhum alerta")
    cab_alerta.font = Font(bold=True, color="FFFFFF", size=10)
    cab_alerta.fill = PatternFill("solid", fgColor="9C0006" if alertas else "276221")
    cab_alerta.alignment = Alignment(horizontal="center")
    ws.row_dimensions[lr].height = 18
    lr += 1

    if alertas:
        _cabecalho(ws, lr, ["Tipo", "Severidade", "Mensagem", "", "", "", "", "", "", "", "", "", "", ""])
        for col in range(3, n_cols + 1):
            ws.merge_cells(f"{get_column_letter(col)}{lr}:{get_column_letter(n_cols)}{lr}")
            break
        ws.row_dimensions[lr].height = 26
        lr += 1
        for alerta in alertas:
            sev = alerta.get("severidade", "INFO")
            cor = COR_ALERTA if sev == "ALERTA" else (COR_AVISO if sev == "AVISO" else COR_LINHA_PAR)
            cf  = COR_ALERTA_FONT if sev == "ALERTA" else (COR_AVISO_FONT if sev == "AVISO" else None)
            _celula(ws, lr, 1, alerta.get("tipo"), cor, True, cf, alinhamento="center")
            _celula(ws, lr, 2, sev, cor, True, cf, alinhamento="center")
            c = ws.cell(row=lr, column=3, value=alerta.get("mensagem"))
            c.font = Font(color=cf or "000000", size=10)
            c.fill = PatternFill("solid", fgColor=cor)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = _BORDA
            ws.merge_cells(f"C{lr}:{get_column_letter(n_cols)}{lr}")
            ws.row_dimensions[lr].height = 30
            lr += 1

    lr += 1  # espaço

    # ── Seção 3: Comparativo de Itens PO vs Cotação vencedora ──
    ws.merge_cells(f"A{lr}:{get_column_letter(n_cols)}{lr}")
    c = ws.cell(row=lr, column=1,
                value=f"COMPARATIVO DE ITENS — PO vs Cotação ({melhor_nome or '—'})")
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor="2E539E")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[lr].height = 18
    lr += 1

    _cabecalho(ws, lr, ["#", "PN (PO)", "Descrição (PO)", "Preço Unit (PO)",
                         "PN (Cotação)", "Descrição (Cotação)", "Preço Unit (Cotação)",
                         "Preço OK?", "PN OK?", "", "", "", "", ""])
    ws.row_dimensions[lr].height = 26
    lr += 1

    itens_melhor = {(it.get("pn") or "").upper(): it for it in melhor.get("itens", []) if it.get("pn")}
    for j, item in enumerate(po.get("itens") or []):
        pn_po   = item.get("pn") or ""
        pn_forn = item.get("pn_fornecedor") or ""
        busca   = (pn_forn or pn_po).upper()
        ref = itens_melhor.get(busca) or next(
            (v for k, v in itens_melhor.items() if busca in k or k in busca), {})
        pn_ok    = "✓" if ref else "VERIFICAR"
        preco_po = item.get("preco_unitario")
        preco_cot = ref.get("preco_unitario") if ref else None
        preco_ok = "✓" if (preco_po and preco_cot and abs(float(preco_po or 0) - float(preco_cot or 0)) < 0.01) else ("VERIFICAR" if ref else "—")
        cor = COR_LINHA_PAR if j % 2 == 0 else COR_BRANCO
        cor_pn = COR_ALERTA if pn_ok == "VERIFICAR" else cor
        cor_pr = COR_AVISO if preco_ok == "VERIFICAR" else cor

        dados = [j+1, pn_po, item.get("descricao"), preco_po,
                 ref.get("pn") if ref else "—", ref.get("descricao") if ref else "—", preco_cot,
                 preco_ok, pn_ok, "", "", "", "", ""]
        fmts = [None,None,None,'"$"#,##0.0000',None,None,'"$"#,##0.0000',None,None,None,None,None,None,None]
        for col, (val, fmt) in enumerate(zip(dados, fmts), 1):
            cf = cor_pn if col == 9 else (cor_pr if col == 8 else cor)
            _celula(ws, lr, col, val, cf, formato=fmt)
        ws.merge_cells(f"J{lr}:{get_column_letter(n_cols)}{lr}")
        ws.row_dimensions[lr].height = 18
        lr += 1

    # Larguras
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    for col in range(10, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


# =====================================================================
# ABA ROBO CONSOLIDADA — todos os pares em uma única aba
# =====================================================================

def _aba_robo_consolidada(wb, lote):
    """
    Cria uma única aba 'ROBO' com todos os itens de todos os pares do lote,
    prontos para copiar e colar diretamente na aba 'pos' do Req-o-matic.
    """
    ws = wb.create_sheet("ROBO")

    # Cabeçalho — cols A-S espelham pos do Req-o-matic; col T = status; col U = Qty; col V = Ship VIA
    colunas = [
        "Quotation Code", "Produto", "Description", "Unit Price",
        "Cost Center Desc", "Supplier", "Freight", "Status", "OBS",
        "ID+QUOTE", "ECO REQ", "Quote+PO", "Coluna3", "Coluna4",
        "Observação PO", "Forn. Extraido", "Forn Extraído ECO", "PO", "Coluna5",
        "Análise",       # col T — status de divergência (NÃO colar no Req-o-matic)
        "Qty (PO)",      # col U — quantidade da PO para conferência com ECO REQ
        "Ship VIA (ECO)",# col V — opção Ship VIA a selecionar no ECO (vermelho = preencher manualmente)
    ]
    for col, titulo in enumerate(colunas, 1):
        cor_cab = "9C0006" if col == 20 else ("FF6600" if col in (21, 22) else "1F3864")
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=cor_cab)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    linha_atual = 2

    for i, entrada in enumerate(lote):
        analise = entrada["analise"]
        po = analise.get("po", {})
        alertas_po = analise.get("alertas_po", [])
        melhor = analise.get("melhor_preco") or {}

        eco_req      = po.get("numero_eco_req") or _req_de_cotacao(analise) or ""
        numero_cot   = _quotation_code(analise)
        numero_po    = po.get("numero_po") or ""
        centro_de_custo = po.get("centro_de_custo") or po.get("solicitante") or ""
        forn_extraido   = po.get("fornecedor_escolhido_comentario") or ""
        observacoes     = po.get("observacoes") or ""
        # Fallback em cadeia: comentários → observações → melhor preço cotação
        if not forn_extraido:
            forn_extraido = _buscar_fornecedor_nas_obs(observacoes)
        if not forn_extraido:
            forn_extraido = melhor.get("nome") or ""
        fornecedor_eco  = _lookup_fornecedor_eco(forn_extraido) or forn_extraido or ""
        freight_robo    = _normalizar_freight_robo(melhor.get("tipo_freight"), melhor.get("custo_freight"))
        obs_alertas     = "; ".join(
            f"[{a.get('tipo')}] {a.get('mensagem', '')[:60]}"
            for a in alertas_po
        ) if alertas_po else ""

        # Índice de itens do fornecedor vencedor para comparação
        itens_melhor = {
            (it.get("pn") or "").upper(): it
            for it in melhor.get("itens", []) if it.get("pn")
        }
        # Índice normalizado (sem hífens/espaços) para match robusto
        itens_melhor_norm = {_norm_pn(k): v for k, v in itens_melhor.items()}

        # Linha separadora entre pares (exceto antes do primeiro)
        tipo_freight    = (melhor.get("tipo_freight") or "").strip()
        ship_via_eco    = SHIP_VIA_MAP.get(tipo_freight.lower())

        if i > 0:
            ws.merge_cells(f"A{linha_atual}:V{linha_atual}")
            sep = ws.cell(row=linha_atual, column=1,
                          value=f"— PO {numero_po}  |  REQ {eco_req}  |  Cotação {numero_cot} —")
            sep.font = Font(bold=True, color="FFFFFF", size=9)
            sep.fill = PatternFill("solid", fgColor="2E539E")
            sep.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[linha_atual].height = 14
            linha_atual += 1

        itens = po.get("itens") or [{}]
        for j, item in enumerate(itens):
            pn_interno = item.get("pn") or ""
            pn_forn    = item.get("pn_fornecedor") or ""
            descricao  = item.get("descricao") or ""
            preco_unit = item.get("preco_unitario")
            quantidade = item.get("quantidade")
            # Vendor por item tem prioridade sobre vendor geral da PO
            forn_item     = (item.get("fornecedor_item") or "").strip()
            fornecedor_eco_item = _lookup_fornecedor_eco(forn_item) or forn_item or fornecedor_eco
            id_quote   = f"{pn_interno}{numero_cot}"
            quote_po   = f"{numero_po}{numero_cot}"
            coluna5    = f"PO:{numero_po} - {numero_cot} - {centro_de_custo}"

            # --- Status de divergência (col T) ---
            busca = (pn_forn or pn_interno).upper()
            busca_norm = _norm_pn(busca)
            ref = (itens_melhor.get(busca)
                   or next((v for k, v in itens_melhor.items()
                             if busca and (busca in k or k in busca)), None)
                   or (itens_melhor_norm.get(busca_norm) if busca_norm else None)
                   or next((v for k, v in itens_melhor_norm.items()
                             if busca_norm and (busca_norm in k or k in busca_norm)), None))
            divergencias = []
            tipos_alerta = {a.get("tipo", "") for a in alertas_po}
            if "PRECO" in tipos_alerta:
                divergencias.append("Preço divergente")
            if "PN" in tipos_alerta or "PART_NUMBER" in tipos_alerta:
                divergencias.append("PN divergente")
            if "FREIGHT" in tipos_alerta:
                divergencias.append("Freight divergente")
            if "FORNECEDOR" in tipos_alerta or "SUPPLIER" in tipos_alerta:
                divergencias.append("Fornecedor divergente")
            if busca and not ref:
                divergencias.append("PN não encontrado na cotação")
            # Preço do item vs cotação
            if ref and preco_unit is not None:
                preco_cot = ref.get("preco_unitario")
                if preco_cot is not None:
                    try:
                        if abs(float(preco_unit) - float(preco_cot)) > 0.01:
                            divergencias.append(f"Preço item: PO={preco_unit} / Cot={preco_cot}")
                    except (TypeError, ValueError):
                        pass
            # Quantidade do item vs cotação vencedora
            if ref and quantidade is not None:
                qtd_cot = ref.get("quantidade")
                if qtd_cot is not None:
                    try:
                        if int(float(quantidade)) != int(float(qtd_cot)):
                            divergencias.append(f"Qty: PO={int(float(quantidade))} / Cot={int(float(qtd_cot))}")
                    except (TypeError, ValueError):
                        pass

            if divergencias:
                status_txt = "⚠ " + " | ".join(divergencias)
                cor_status = COR_ALERTA
                cor_font_status = COR_ALERTA_FONT
                cor_linha = COR_ALERTA
            else:
                status_txt = "✓ OK"
                cor_status = COR_MELHOR
                cor_font_status = COR_MELHOR_FONT
                cor_linha = COR_LINHA_PAR if j % 2 == 0 else COR_BRANCO

            dados = [
                numero_cot, pn_interno, descricao, preco_unit,
                centro_de_custo, fornecedor_eco_item, freight_robo,
                "", obs_alertas, id_quote, eco_req, quote_po,
                "", "", observacoes, forn_item or forn_extraido, "", numero_po, coluna5,
            ]

            for col, val in enumerate(dados, 1):
                cf = COR_AVISO if col == 9 and obs_alertas else cor_linha
                fmt = '"$"#,##0.0000' if col == 4 else None
                _celula(ws, linha_atual, col, val, cf, formato=fmt)

            # Col T — status de análise
            _celula(ws, linha_atual, 20, status_txt,
                    cor_status, negrito=True, cor_fonte=cor_font_status)

            # Col U — quantidade da PO (laranja se divergente)
            cor_qty = COR_ALERTA if any("Qty" in d for d in divergencias) else cor_linha
            _celula(ws, linha_atual, 21, quantidade, cor_qty, alinhamento="center")

            # Col V — Ship VIA (verde se mapeado, vermelho se precisar preencher manualmente)
            if ship_via_eco:
                _celula(ws, linha_atual, 22, ship_via_eco, COR_MELHOR,
                        cor_fonte=COR_MELHOR_FONT, alinhamento="center")
            else:
                txt_manual = f"(manual — {tipo_freight})" if tipo_freight else "(manual)"
                _celula(ws, linha_atual, 22, txt_manual, COR_ALERTA,
                        negrito=True, cor_fonte=COR_ALERTA_FONT, alinhamento="center")

            ws.row_dimensions[linha_atual].height = 18
            linha_atual += 1

    # Larguras fixas espelhando a aba pos do Req-o-matic + cols T e U
    larguras = {
        "A": 16, "B": 14, "C": 45, "D": 14, "E": 20, "F": 28,
        "G": 16, "H": 14, "I": 35, "J": 28, "K": 18, "L": 28,
        "M": 10, "N": 10, "O": 45, "P": 25, "Q": 25, "R": 12, "S": 40,
        "T": 35,   # Análise
        "U": 10,   # Qty (PO)
        "V": 22,   # Ship VIA (ECO)
    }
    for col_letter, w in larguras.items():
        ws.column_dimensions[col_letter].width = w


# =====================================================================
# Função principal — arquivo único para o lote
# =====================================================================

def exportar_excel(lote: list, pasta_saida: str) -> str:
    """
    lote: lista de dicts com chaves 'analise' e 'req_numero' (opcional)
    Gera um único arquivo Excel com todas as análises.
    Retorna o caminho do arquivo gerado.
    """
    wb = Workbook()
    wb.remove(wb.active)

    _aba_indice(wb, lote)

    for i, entrada in enumerate(lote):
        prefixo = str(i + 1)
        analise = entrada["analise"]
        _aba_analise(wb, analise, prefixo)

    # Aba ROBO consolidada — todos os pares juntos para copiar no Req-o-matic
    _aba_robo_consolidada(wb, lote)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = f"Analise_Cotacoes_{timestamp}.xlsx"
    caminho = os.path.join(pasta_saida, nome)
    wb.save(caminho)
    return caminho
