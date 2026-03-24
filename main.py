"""
main.py
-------
Interface gráfica principal do sistema de análise de cotações e PO.
Execute este arquivo com duplo clique ou: python main.py
"""

import os
import re
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Verifica se as dependências estão instaladas
try:
    import google.genai
    import openpyxl
    import pdfplumber
except ImportError as e:
    import subprocess
    resposta = tk.messagebox.askyesno(
        "Dependências não instaladas",
        f"Módulo ausente: {e}\n\nDeseja instalar as dependências agora?\n"
        "(Requer conexão com a internet)"
    )
    if resposta:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])
        messagebox.showinfo("Instalação concluída", "Dependências instaladas. Reinicie o programa.")
    sys.exit(0)

from config import GOOGLE_API_KEY
from extractor import extrair_cotacoes, extrair_po
from analyzer import analisar
from excel_exporter import exportar_excel


# =====================================================================
# Helpers
# =====================================================================

def _extrair_req_do_nome(caminho: str):
    """Extrai o Nº ECO REQ do nome do arquivo. Ex: 'Cotação REQ 031326015461' → '031326015461'"""
    nome = os.path.basename(caminho)
    match = re.search(r'REQ\s*(\d{8,})', nome, re.IGNORECASE)
    return match.group(1) if match else None


def _e_po(caminho: str) -> bool:
    """
    Heurística: arquivo é PO se o nome indicar uma Purchase Order.
    Cobre padrões como: 'PO 031326015461', 'PO BRAM - 430581', 'PO 475919 - QT...',
    'Purchase Order', '467694 - ECO' (gerado pelo ECO Requisition).
    """
    nome = os.path.basename(caminho).lower()
    # Começa com "po " ou "po_" ou "po-" (ex: "PO 475919 - QT 2026.007638")
    if re.match(r'^po[\s\-_]', nome):
        return True
    # Outros padrões comuns
    return any(p in nome for p in (
        "purchase order", "ordem de compra",
        "- eco", "_eco",  # ex: "467694 - ECO"
    ))


def _parear_por_req(cotacoes: list, pos: list) -> tuple:
    """
    Pareia cotações com POs pelo nº REQ no nome do arquivo.
    Retorna (pares, nao_pareados_cotacoes, nao_pareados_pos).
    """
    idx_cot = {}  # req → caminho
    for c in cotacoes:
        req = _extrair_req_do_nome(c)
        if req:
            idx_cot[req] = c
        else:
            idx_cot[f"__sem_req_{c}"] = c

    idx_po = {}
    for p in pos:
        req = _extrair_req_do_nome(p)
        if req:
            idx_po[req] = p
        else:
            idx_po[f"__sem_req_{p}"] = p

    pares = []
    usados_cot = set()
    usados_po = set()

    # Pareia por REQ
    for req, cot in idx_cot.items():
        if req in idx_po:
            pares.append((cot, idx_po[req]))
            usados_cot.add(cot)
            usados_po.add(idx_po[req])

    # Arquivos sem REQ no nome: pareia por ordem
    sem_req_cot = [c for c in cotacoes if c not in usados_cot]
    sem_req_po  = [p for p in pos       if p not in usados_po]
    for cot, po in zip(sem_req_cot, sem_req_po):
        pares.append((cot, po))
        usados_cot.add(cot)
        usados_po.add(po)

    nao_cot = [c for c in cotacoes if c not in usados_cot]
    nao_po  = [p for p in pos       if p not in usados_po]
    return pares, nao_cot, nao_po


# =====================================================================
# Verifica configuração da API Key
# =====================================================================

def _verificar_api_key():
    chave_invalida = (
        "COLE-SUA-CHAVE-AQUI" in GOOGLE_API_KEY
        or len(GOOGLE_API_KEY) < 10
        or not GOOGLE_API_KEY.startswith("AIzaSy")
    )
    if chave_invalida:
        messagebox.showwarning(
            "API Key inválida",
            "A chave da API Google Gemini está ausente ou incorreta.\n\n"
            "A chave deve começar com 'AIzaSy...' (39 caracteres).\n\n"
            "Como obter:\n"
            "1. Acesse aistudio.google.com e faça login\n"
            "2. Clique em 'Get API key' → 'Create API key'\n"
            "3. Copie a chave e cole em config.py\n"
            "4. Salve o arquivo e reinicie o programa"
        )
        return False
    return True


# =====================================================================
# Diálogo de credenciais ECO Requisition
# =====================================================================

class DialogCredenciais(tk.Toplevel):
    """Solicita usuário e senha do ECO Requisition."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Login — ECO Requisition")
        self.resizable(False, False)
        self.configure(bg="white")
        self.grab_set()
        self.resultado = None
        self._usuario = tk.StringVar()
        self._senha   = tk.StringVar()
        self._construir()
        self._centralizar(parent)

    def _centralizar(self, parent):
        self.update_idletasks()
        px = parent.winfo_x() + parent.winfo_width()  // 2
        py = parent.winfo_y() + parent.winfo_height() // 2
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{px - w // 2}+{py - h // 2}")

    def _construir(self):
        pad = {"padx": 24, "pady": 6}

        tk.Label(self, text="Credenciais do ECO Requisition",
                 font=("Segoe UI", 11, "bold"), bg="white",
                 fg="#1F3864").pack(pady=(20, 4))
        tk.Label(self, text="Os dados são usados apenas nesta sessão e não são salvos.",
                 font=("Segoe UI", 8), bg="white", fg="#888").pack(pady=(0, 12))

        for label, var, show in [("Usuário:", self._usuario, ""),
                                  ("Senha:",   self._senha,   "•")]:
            tk.Label(self, text=label, font=("Segoe UI", 9),
                     bg="white", anchor="w").pack(fill="x", **pad)
            tk.Entry(self, textvariable=var, show=show,
                     font=("Segoe UI", 10), relief="solid", bd=1,
                     width=32).pack(padx=24, pady=(0, 6), ipady=4)

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=24, pady=12)

        frame = tk.Frame(self, bg="white")
        frame.pack(pady=(0, 18))
        tk.Button(frame, text="Cancelar", command=self.destroy,
                  font=("Segoe UI", 9), bg="#E8E8E8", relief="flat",
                  cursor="hand2", padx=16, pady=6).pack(side="left", padx=6)
        tk.Button(frame, text="✓  Entrar e criar POs", command=self._confirmar,
                  font=("Segoe UI", 9, "bold"), bg="#276221", fg="white",
                  activebackground="#1a4016", activeforeground="white",
                  relief="flat", cursor="hand2", padx=16, pady=6).pack(side="left", padx=6)

        self.bind("<Return>", lambda _: self._confirmar())

    def _confirmar(self):
        u = self._usuario.get().strip()
        s = self._senha.get()
        if not u or not s:
            messagebox.showwarning("Campos obrigatórios",
                                   "Informe usuário e senha.", parent=self)
            return
        self.resultado = (u, s)
        self.destroy()


# =====================================================================
# Janela principal
# =====================================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Análise de Cotações e PO — ECO Purchasing")
        self.resizable(False, False)
        self.configure(bg="#1F3864")

        self._fila = []   # Lista de (caminho_cotacao, caminho_po)
        self._lote = []   # Resultado da análise — alimenta o Playwright
        self._pasta_saida = tk.StringVar(value=os.path.dirname(os.path.abspath(__file__)))

        self._construir_ui()
        self._centralizar()

    def _centralizar(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    def _construir_ui(self):
        # --- Cabeçalho ---
        header = tk.Frame(self, bg="#1F3864", pady=16, padx=24)
        header.pack(fill="x")
        tk.Label(header, text="Análise de Cotações & PO",
                 font=("Segoe UI", 18, "bold"), fg="white", bg="#1F3864").pack()
        tk.Label(header, text="ECO Purchasing  |  Powered by Gemini AI",
                 font=("Segoe UI", 9), fg="#A0B4D0", bg="#1F3864").pack()

        # --- Corpo ---
        corpo = tk.Frame(self, bg="white", padx=28, pady=20)
        corpo.pack(fill="both", expand=True)

        # Botões de importação (topo)
        frame_import = tk.Frame(corpo, bg="#F0F4FA", bd=1, relief="solid")
        frame_import.pack(fill="x", pady=(0, 12))

        tk.Label(frame_import, text="Adicionar arquivos:",
                 font=("Segoe UI", 9, "bold"), bg="#F0F4FA", fg="#1F3864"
                 ).pack(side="left", padx=(12, 8), pady=8)

        tk.Button(frame_import, text="📂  Importar pasta",
                  command=self._importar_pasta,
                  font=("Segoe UI", 9), bg="#1F3864", fg="white",
                  activebackground="#2E539E", activeforeground="white",
                  relief="flat", cursor="hand2", padx=12, pady=5
                  ).pack(side="left", padx=4, pady=6)

        tk.Button(frame_import, text="📄  Selecionar múltiplas Cotações + POs",
                  command=self._importar_multiplos,
                  font=("Segoe UI", 9), bg="#2E539E", fg="white",
                  activebackground="#1F3864", activeforeground="white",
                  relief="flat", cursor="hand2", padx=12, pady=5
                  ).pack(side="left", padx=4, pady=6)

        ttk.Separator(frame_import, orient="vertical").pack(side="left", fill="y", pady=6)

        tk.Button(frame_import, text="📊  Carregar Excel ROBO",
                  command=self._carregar_excel_robo,
                  font=("Segoe UI", 9), bg="#276221", fg="white",
                  activebackground="#1a4016", activeforeground="white",
                  relief="flat", cursor="hand2", padx=12, pady=5
                  ).pack(side="left", padx=4, pady=6)

        tk.Label(frame_import,
                 text="Pareamento automático por nº REQ no nome do arquivo",
                 font=("Segoe UI", 8), bg="#F0F4FA", fg="#666"
                 ).pack(side="left", padx=(8, 12), pady=6)

        # Label fila
        tk.Label(corpo, text="Fila de análises:",
                 font=("Segoe UI", 10, "bold"), bg="white", anchor="w"
                 ).pack(fill="x", pady=(0, 4))

        # Cabeçalho da tabela
        cab = tk.Frame(corpo, bg="#1F3864")
        cab.pack(fill="x")
        tk.Label(cab, text="#",    width=3,  font=("Segoe UI", 8, "bold"),
                 fg="white", bg="#1F3864", anchor="center").pack(side="left", padx=(4, 0), pady=3)
        tk.Label(cab, text="Cotações (PDF)", width=30, font=("Segoe UI", 8, "bold"),
                 fg="white", bg="#1F3864", anchor="w").pack(side="left", padx=4)
        tk.Label(cab, text="PO (PDF)", width=30, font=("Segoe UI", 8, "bold"),
                 fg="white", bg="#1F3864", anchor="w").pack(side="left", padx=4)

        # Lista com scrollbar
        frame_scroll = tk.Frame(corpo, bg="#EEEEEE", bd=1, relief="solid")
        frame_scroll.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame_scroll, orient="vertical")
        self._listbox = tk.Listbox(
            frame_scroll, yscrollcommand=scrollbar.set,
            font=("Segoe UI", 9), height=7,
            selectbackground="#2E539E", selectforeground="white",
            activestyle="none", bg="white", relief="flat",
            highlightthickness=0
        )
        scrollbar.config(command=self._listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self._listbox.pack(side="left", fill="both", expand=True)

        # Botão remover
        frame_fila_btns = tk.Frame(corpo, bg="white")
        frame_fila_btns.pack(fill="x", pady=(4, 0))
        tk.Button(frame_fila_btns, text="🗑  Remover selecionado",
                  command=self._remover_par,
                  font=("Segoe UI", 9), bg="#FFE8E8", fg="#9C0006",
                  relief="flat", cursor="hand2", padx=10, pady=3
                  ).pack(side="left")
        tk.Button(frame_fila_btns, text="✖  Limpar tudo",
                  command=self._limpar_fila,
                  font=("Segoe UI", 9), bg="#FFE8E8", fg="#9C0006",
                  relief="flat", cursor="hand2", padx=10, pady=3
                  ).pack(side="left", padx=(6, 0))

        ttk.Separator(corpo, orient="horizontal").pack(fill="x", pady=12)

        # Pasta de saída
        tk.Label(corpo, text="Salvar Excel em:",
                 font=("Segoe UI", 10), bg="white", anchor="w"
                 ).pack(fill="x", pady=(0, 4))
        frame_saida = tk.Frame(corpo, bg="white")
        frame_saida.pack(fill="x", pady=(0, 14))
        tk.Entry(frame_saida, textvariable=self._pasta_saida,
                 font=("Segoe UI", 9), fg="#333", relief="solid", bd=1,
                 state="readonly", readonlybackground="#F5F5F5", width=56
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(frame_saida, text="Pasta...",
                  command=self._selecionar_pasta_saida,
                  font=("Segoe UI", 9), bg="#E8E8E8", relief="flat",
                  cursor="hand2", padx=8
                  ).pack(side="left", padx=(6, 0))

        # Barra de progresso
        self._progress = ttk.Progressbar(corpo, mode="determinate", length=520)
        self._progress.pack(fill="x", pady=(0, 4))

        # Status
        self._status = tk.StringVar(value="Importe uma pasta ou selecione os PDFs e clique em Analisar.")
        tk.Label(corpo, textvariable=self._status,
                 font=("Segoe UI", 9), bg="white", fg="#555",
                 wraplength=520, justify="left", anchor="w"
                 ).pack(fill="x", pady=(0, 12))

        # Botões de ação
        frame_acoes = tk.Frame(corpo, bg="white")
        frame_acoes.pack(pady=(0, 4))

        self._btn_analisar = tk.Button(
            frame_acoes, text="  ▶  ANALISAR TODOS",
            command=self._iniciar_analise,
            font=("Segoe UI", 12, "bold"),
            bg="#1F3864", fg="white",
            activebackground="#2E539E", activeforeground="white",
            relief="flat", cursor="hand2", pady=10, padx=24,
        )
        self._btn_analisar.pack(side="left", padx=(0, 8))

        self._btn_criar_pos = tk.Button(
            frame_acoes, text="  🌐  CRIAR POs NO ECO",
            command=self._iniciar_criacao_pos,
            font=("Segoe UI", 12, "bold"),
            bg="#276221", fg="white",
            activebackground="#1a4016", activeforeground="white",
            relief="flat", cursor="hand2", pady=10, padx=24,
            state="disabled",   # habilitado após análise concluída
        )
        self._btn_criar_pos.pack(side="left")

    # ------------------------------------------------------------------
    # Importação de arquivos
    # ------------------------------------------------------------------

    def _importar_pasta(self):
        """
        Seleciona uma pasta e pareia os PDFs em sequência por data de modificação.
        Os arquivos são salvos em pares consecutivos (PO + cotação ou cotação + PO),
        então basta ordenar por data e parear de dois em dois.
        """
        pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs de Cotações e POs")
        if not pasta:
            return

        todos = [
            os.path.join(pasta, f)
            for f in os.listdir(pasta)
            if f.lower().endswith(".pdf")
        ]
        if not todos:
            messagebox.showinfo("Nenhum PDF", "Nenhum arquivo PDF encontrado na pasta selecionada.")
            return

        # Ordena por data de modificação (mais antigo primeiro = mesma ordem de salvamento)
        todos.sort(key=lambda f: os.path.getmtime(f))

        pares = []
        nao_pareados = []

        i = 0
        while i + 1 < len(todos):
            f1, f2 = todos[i], todos[i + 1]
            e_po_f1 = _e_po(f1)
            e_po_f2 = _e_po(f2)

            if e_po_f1 and not e_po_f2:
                pares.append((f2, f1))   # (cotação, PO)
            elif e_po_f2 and not e_po_f1:
                pares.append((f1, f2))   # (cotação, PO)
            else:
                # Não foi possível distinguir — assume f1=cotação, f2=PO por ordem
                pares.append((f1, f2))
            i += 2

        # Se total ímpar, sobra um arquivo sem par
        if len(todos) % 2 != 0:
            nao_pareados.append(os.path.basename(todos[-1]))

        if not pares:
            messagebox.showwarning("Sem pares", "Nenhum par pôde ser formado com os PDFs da pasta.")
            return

        for cot, po in pares:
            self._fila.append((cot, po))
            nome_cot = os.path.basename(cot)[:32]
            nome_po  = os.path.basename(po)[:32]
            n = len(self._fila)
            self._listbox.insert("end", f"  {n}   {nome_cot:<32}  {nome_po}")

        msg = f"{len(self._fila)} par(es) na fila."
        if nao_pareados:
            msg += f"  ⚠ Arquivo sem par: {nao_pareados[0]}"
            messagebox.showwarning("Arquivo sem par",
                                   f"O arquivo '{nao_pareados[0]}' ficou sem par.\n"
                                   "Total de PDFs na pasta é ímpar.")
        self._status.set(msg)

    def _importar_multiplos(self):
        """
        Seleção manual em duas etapas:
        1) Seleciona N PDFs de cotações (multi-select)
        2) Seleciona N PDFs de POs     (multi-select)
        Pareia automaticamente por nº REQ ou por ordem.
        """
        cotacoes = filedialog.askopenfilenames(
            title="Selecione os PDFs de COTAÇÕES (pode selecionar vários)",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")]
        )
        if not cotacoes:
            return

        pos = filedialog.askopenfilenames(
            title="Selecione os PDFs de PO (pode selecionar vários)",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")]
        )
        if not pos:
            return

        self._adicionar_pares(list(cotacoes), list(pos), origem="seleção manual")

    def _adicionar_pares(self, cotacoes: list, pos: list, origem: str = ""):
        """Para a lista de cotações e POs, pareia e adiciona à fila."""
        pares, sem_cot, sem_po = _parear_por_req(cotacoes, pos)

        if not pares:
            messagebox.showwarning(
                "Sem pares",
                "Nenhum par Cotação+PO pôde ser formado.\n\n"
                "Verifique se os arquivos têm o mesmo nº REQ no nome."
            )
            return

        for cot, po in pares:
            self._fila.append((cot, po))
            nome_cot = os.path.basename(cot)[:32]
            nome_po  = os.path.basename(po)[:32]
            n = len(self._fila)
            self._listbox.insert("end", f"  {n}   {nome_cot:<32}  {nome_po}")

        msg = f"{len(self._fila)} par(es) na fila."
        avisos = []
        if sem_cot:
            avisos.append(f"{len(sem_cot)} cotação(ões) sem PO correspondente.")
        if sem_po:
            avisos.append(f"{len(sem_po)} PO(s) sem cotação correspondente.")
        if avisos:
            msg += "  ⚠ " + "  ".join(avisos)
        self._status.set(msg)

        if avisos:
            messagebox.showwarning(
                "Arquivos não pareados",
                "\n".join(avisos) + "\n\nVerifique se o nº REQ aparece no nome dos arquivos."
            )

    def _remover_par(self):
        sel = self._listbox.curselection()
        if not sel:
            messagebox.showinfo("Nenhum selecionado", "Clique em um item da lista para selecioná-lo.")
            return
        idx = sel[0]
        self._listbox.delete(idx)
        self._fila.pop(idx)
        self._renumerar_lista()
        self._status.set(f"{len(self._fila)} par(es) na fila.")

    def _limpar_fila(self):
        if not self._fila:
            return
        if messagebox.askyesno("Limpar fila", "Remover todos os pares da fila?"):
            self._fila.clear()
            self._listbox.delete(0, "end")
            self._status.set("Fila limpa. Importe os arquivos para começar.")

    def _renumerar_lista(self):
        itens = list(self._listbox.get(0, "end"))
        self._listbox.delete(0, "end")
        for i, item in enumerate(itens):
            partes = item.split("   ", 1)
            novo = f"  {i + 1}   {partes[1] if len(partes) > 1 else item}"
            self._listbox.insert("end", novo)

    def _carregar_excel_robo(self):
        """
        Carrega a aba ROBO de um Excel gerado anteriormente e reconstrói
        self._lote no formato esperado pelo eco_playwright, habilitando
        o botão 'CRIAR POs NO ECO' sem precisar re-executar a análise.
        """
        caminho = filedialog.askopenfilename(
            title="Selecione o arquivo Excel com a aba ROBO",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
        )
        if not caminho:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(caminho, read_only=True, data_only=True)

            if "ROBO" not in wb.sheetnames:
                messagebox.showerror(
                    "Aba não encontrada",
                    "O arquivo selecionado não contém a aba 'ROBO'.\n\n"
                    "Selecione um arquivo gerado por este programa."
                )
                wb.close()
                return

            ws = wb["ROBO"]
            # Colunas (1-based → índice 0-based):
            # A(0)=Quotation Code  B(1)=Produto  C(2)=Description  D(3)=Unit Price
            # E(4)=Cost Center     F(5)=Supplier  K(10)=ECO REQ
            # O(14)=Observação PO  P(15)=Forn Extraido  R(17)=PO
            grupos = {}      # numero_po → dict com dados acumulados
            ordem_pos = []   # preserva ordem de aparição

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Linha separadora: col A começa com "—" ou col R está vazia
                col_a = str(row[0] or "")
                if col_a.startswith("—"):
                    continue

                numero_po = str(row[17] or "").strip()   # col R
                if not numero_po:
                    continue

                numero_cot   = str(row[0]  or "").strip()   # col A
                pn_interno   = str(row[1]  or "").strip()   # col B
                descricao    = str(row[2]  or "").strip()   # col C
                preco_unit   = row[3]                        # col D (numérico)
                centro_custo = str(row[4]  or "").strip()   # col E
                fornec_eco   = str(row[5]  or "").strip()   # col F
                eco_req      = str(row[10] or "").strip()   # col K
                observacoes  = str(row[14] or "").strip()   # col O
                forn_extraid = str(row[15] or "").strip()   # col P
                # col V (índice 21) = Ship VIA preenchido manualmente
                try:
                    ship_via = str(row[21] or "").strip() if len(row) > 21 else ""
                    # Remove prefixo "(manual)" se presente
                    if ship_via.startswith("(manual"):
                        ship_via = ""
                except (IndexError, TypeError):
                    ship_via = ""

                if numero_po not in grupos:
                    grupos[numero_po] = {
                        "numero_po":    numero_po,
                        "numero_cot":   numero_cot,
                        "eco_req":      eco_req,
                        "centro_custo": centro_custo,
                        "fornec_eco":   fornec_eco,
                        "forn_extraid": forn_extraid,
                        "observacoes":  observacoes,
                        "ship_via":     ship_via,
                        "itens":        [],
                    }
                    ordem_pos.append(numero_po)

                if pn_interno or descricao:
                    try:
                        preco = float(preco_unit) if preco_unit is not None else None
                    except (TypeError, ValueError):
                        preco = None
                    # col U (índice 20) = Qty (PO) — pode não existir em planilhas antigas
                    try:
                        qtd_raw = row[20] if len(row) > 20 else None
                        qtd = int(float(qtd_raw)) if qtd_raw is not None else None
                    except (TypeError, ValueError, IndexError):
                        qtd = None
                    grupos[numero_po]["itens"].append({
                        "pn":            pn_interno,
                        "descricao":     descricao,
                        "preco_unitario": preco,
                        "quantidade":    qtd,
                    })

            wb.close()

            if not grupos:
                messagebox.showwarning(
                    "Nenhum dado",
                    "A aba ROBO não contém linhas com número de PO (col R).\n"
                    "Verifique se o arquivo correto foi selecionado."
                )
                return

            # Reconstrói self._lote no formato que eco_playwright espera
            lote = []
            for numero_po in ordem_pos:
                g = grupos[numero_po]
                lote.append({
                    "analise": {
                        "po": {
                            "numero_po":                    g["numero_po"],
                            "numero_eco_req":               g["eco_req"] or None,
                            "centro_de_custo":              g["centro_custo"] or None,
                            "fornecedor_escolhido_comentario": g["forn_extraid"] or None,
                            "observacoes":                  g["observacoes"] or None,
                            "itens":                        g["itens"],
                        },
                        "melhor_preco": {
                            "nome":          g["fornec_eco"] or None,
                            "ship_via_direto": g["ship_via"] or None,
                        },
                        "ranking_preco": (
                            [{"numero_cotacao": g["numero_cot"]}]
                            if g["numero_cot"] else []
                        ),
                        "alertas_po": [],
                    },
                    "req_numero": g["eco_req"] or None,
                })

            self._lote = lote
            self._btn_criar_pos.config(state="normal")
            n = len(lote)
            self._status.set(
                f"✓ {n} PO(s) carregada(s) do Excel. "
                "Clique em 'CRIAR POs NO ECO' para prosseguir."
            )
            resumo = "\n".join(
                f"  • PO {g['numero_po']}  |  Cot {g['numero_cot'] or '—'}"
                f"  |  {g['fornec_eco'] or '—'}  |  {len(g['itens'])} item(ns)"
                for g in [grupos[p] for p in ordem_pos]
            )
            messagebox.showinfo(
                "Excel ROBO carregado",
                f"{n} PO(s) prontas para criação no ECO:\n\n{resumo}"
            )

        except Exception as exc:
            messagebox.showerror("Erro ao carregar Excel", str(exc))

    def _selecionar_pasta_saida(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta para salvar o Excel")
        if pasta:
            self._pasta_saida.set(pasta)

    # ------------------------------------------------------------------
    # Análise
    # ------------------------------------------------------------------

    def _iniciar_analise(self):
        if not _verificar_api_key():
            return
        if not self._fila:
            messagebox.showwarning("Fila vazia", "Importe pelo menos um par de Cotação + PO.")
            return

        self._btn_analisar.config(state="disabled")
        self._progress["value"] = 0
        self._progress["maximum"] = len(self._fila)
        self._status.set(f"Iniciando análise de {len(self._fila)} par(es)...")

        thread = threading.Thread(target=self._executar_analise, daemon=True)
        thread.start()

    def _executar_analise(self):
        lote = []
        erros = []

        for i, (caminho_cotacao, caminho_po) in enumerate(self._fila):
            nome = f"Par {i + 1}/{len(self._fila)}"
            try:
                self._atualizar_status(f"[{nome}] Extraindo cotações...")
                dados_cotacao = extrair_cotacoes(caminho_cotacao)

                self._atualizar_status(f"[{nome}] Extraindo PO...")
                dados_po = extrair_po(caminho_po)

                self._atualizar_status(f"[{nome}] Comparando e validando...")
                analise = analisar(dados_cotacao, dados_po)

                req_numero = _extrair_req_do_nome(caminho_cotacao) or _extrair_req_do_nome(caminho_po)
                lote.append({"analise": analise, "req_numero": req_numero})

            except Exception as e:
                erros.append((i + 1, str(e)))

            self.after(0, lambda v=i + 1: self._progress.configure(value=v))

        if lote:
            self._atualizar_status("Gerando arquivo Excel unificado...")
            try:
                caminho_excel = exportar_excel(lote, self._pasta_saida.get())
            except Exception as e:
                erros.append(("Excel", str(e)))
                caminho_excel = None
        else:
            caminho_excel = None

        self.after(0, lambda: self._concluir(caminho_excel, lote, erros))

    def _atualizar_status(self, mensagem: str):
        self.after(0, lambda: self._status.set(mensagem))

    def _concluir(self, caminho_excel, lote: list, erros: list):
        self._progress["value"] = self._progress["maximum"]
        self._btn_analisar.config(state="normal")

        n_ok = len(lote)
        n_err = len(erros)
        linhas = [f"✓ Análise concluída — {n_ok} par(es) processado(s), {n_err} erro(s).\n"]

        for i, entrada in enumerate(lote):
            analise = entrada["analise"]
            po = analise.get("po", {})
            alertas = len(analise.get("alertas_po", []))
            melhor = (analise.get("melhor_preco") or {}).get("nome", "N/A")
            numero_po = po.get("numero_po") or f"Par {i + 1}"
            eco_req = po.get("numero_eco_req") or entrada.get("req_numero") or "—"
            linhas.append(
                f"  PO: {numero_po} | REQ: {eco_req} | {alertas} alerta(s) | Melhor: {melhor}"
            )

        for n, erro in erros:
            linhas.append(f"  Par {n}: ERRO — {erro[:80]}")

        if caminho_excel:
            linhas.append(f"\nArquivo salvo: {os.path.basename(caminho_excel)}")
            self._status.set(f"✓ Concluído! {n_ok} par(es), {n_err} erro(s). Excel salvo.")
        else:
            self._status.set(f"Concluído com erros. {n_err} erro(s).")

        # Habilita botão de criação de POs se houver análises OK
        if lote:
            self._lote = lote
            self._btn_criar_pos.config(state="normal")

        resposta = messagebox.askyesno(
            "Análise concluída",
            "\n".join(linhas) + "\n\nDeseja abrir o arquivo Excel?"
        )
        if resposta and caminho_excel:
            os.startfile(caminho_excel)

    def _erro(self, mensagem: str):
        self._btn_analisar.config(state="normal")
        self._status.set(f"Erro: {mensagem}")
        messagebox.showerror("Erro na análise", mensagem)

    # ------------------------------------------------------------------
    # Criação de POs no ECO Requisition via Playwright
    # ------------------------------------------------------------------

    def _iniciar_criacao_pos(self):
        if not self._lote:
            messagebox.showwarning("Sem análise", "Execute a análise primeiro.")
            return

        dialogo = DialogCredenciais(self)
        self.wait_window(dialogo)
        if not dialogo.resultado:
            return

        usuario, senha = dialogo.resultado
        self._btn_criar_pos.config(state="disabled")
        self._btn_analisar.config(state="disabled")
        self._status.set("Iniciando automação do ECO Requisition...")

        thread = threading.Thread(
            target=self._executar_criacao_pos,
            args=(usuario, senha),
            daemon=True
        )
        thread.start()

    def _executar_criacao_pos(self, usuario: str, senha: str):
        from eco_playwright import criar_pos
        import threading
        from tkinter import simpledialog

        def _bloquear(fn):
            """Executa fn no thread principal e aguarda o resultado."""
            evento = threading.Event()
            resultado = [None]
            def _wrap():
                resultado[0] = fn()
                evento.set()
            self.after(0, _wrap)
            evento.wait()
            return resultado[0]

        def confirmar(titulo: str, mensagem: str) -> bool:
            return _bloquear(lambda: messagebox.askyesno(titulo, mensagem))

        def escolher(titulo: str, opcoes: list):
            """
            Exibe as opções numeradas e pede ao usuário que digite o número
            OU um nome de fornecedor livre.
            Retorna: int (índice 0-based), str (nome digitado) ou -1 (cancelou).
            """
            lista = "\n".join(f"  [{i+1}] {op}" for i, op in enumerate(opcoes))
            msg   = (f"{lista}\n\n"
                     "Digite o NÚMERO da opção desejada\n"
                     "ou ESCREVA o nome correto do fornecedor:")
            resp = _bloquear(lambda: simpledialog.askstring(titulo, msg))
            if resp is None:
                return -1
            resp = resp.strip()
            if resp.isdigit():
                n = int(resp)
                if 1 <= n <= len(opcoes):
                    return n - 1
            # Texto livre — retorna como string para o Playwright digitar
            return resp if resp else -1

        try:
            resultados = criar_pos(
                usuario, senha, self._lote,
                callback=lambda msg: self.after(0, lambda m=msg: self._status.set(m)),
                confirmar=confirmar,   # remover para modo autônomo
                escolher=escolher,     # remover para modo autônomo (usa vendor_map salvo)
            )
        except Exception as e:
            resultados = [{"po": "?", "status": "ERRO", "mensagem": str(e)}]
        self.after(0, lambda: self._concluir_pos(resultados))

    def _concluir_pos(self, resultados: list):
        self._btn_criar_pos.config(state="normal")
        self._btn_analisar.config(state="normal")

        ok    = [r for r in resultados if r.get("status") == "OK"]
        erros = [r for r in resultados if r.get("status") != "OK"]

        from datetime import datetime
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")
        linhas_log = [
            f"LOG DE EMISSÃO DE POs — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Total: {len(resultados)} | OK: {len(ok)} | Erros: {len(erros)}",
            "=" * 60,
        ]
        for r in resultados:
            icone = "OK" if r.get("status") == "OK" else "ERRO"
            po_num = r.get("po", "?")
            msg    = r.get("mensagem", "")
            po_gerado = r.get("po_gerado", "")
            linha = f"[{icone}] PO {po_num}"
            if po_gerado and po_gerado != po_num:
                linha += f" → {po_gerado}"
            if msg:
                linha += f" | {msg}"
            linhas_log.append(linha)

        # Grava o log na pasta do projeto
        pasta_log = os.path.dirname(os.path.abspath(__file__))
        caminho_log = os.path.join(pasta_log, f"Log_POs_{agora}.txt")
        try:
            with open(caminho_log, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas_log))
            os.startfile(caminho_log)   # abre automaticamente no Bloco de Notas
        except Exception:
            pass  # se falhar ao abrir, não bloqueia

        self._status.set(f"ECO: {len(ok)} PO(s) criada(s), {len(erros)} erro(s). Log salvo.")
        linhas_resumo = [f"✓ Criação de POs concluída — {len(ok)} OK, {len(erros)} erro(s).\n"]
        for r in resultados:
            icone = "✓" if r.get("status") == "OK" else "✗"
            linhas_resumo.append(f"  {icone}  PO {r.get('po')}: {r.get('mensagem')}")
        messagebox.showinfo("Criação de POs", "\n".join(linhas_resumo))


# =====================================================================
# Ponto de entrada
# =====================================================================

if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    app = App()
    app.mainloop()
